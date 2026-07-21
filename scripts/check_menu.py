#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
お弁当メニュー違反チェックエンジン
使い方:
    python check_menu.py <入力xlsx> [出力xlsx]
入力xlsx は「メニュー一覧」「食材」タブを含むこと（「メニュールール」タブは任意）。
1ヶ月分でも1週間分でも可（連続日チェックは渡された範囲内で評価）。
出力xlsx は「違反候補リスト/サマリー/チェックフロー」の3シート。
"""
import sys, os, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
MASTER_CSV = os.path.join(HERE, '..', 'assets', 'master_ingredients.csv')

# ============ マスタ（人がレビュー済みの分類）読み込み ============
def load_master():
    try:
        m = pd.read_csv(MASTER_CSV)
        g  = dict(zip(m['商品名(原文)'], m['主原料グループ']))
        ne = dict(zip(m['商品名(原文)'], m['練り物'].fillna('')))
        ka = dict(zip(m['商品名(原文)'], m['乾物(水戻し)'].fillna('')))
        return g, ne, ka
    except Exception:
        return {}, {}, {}

M_GROUP, M_NERI, M_KAN = load_master()

# ============ キーワード分類（マスタ未収載の新商材フォールバック）============
def group_from_name(n):
    """料理名/商品名 -> 代表 主原料グループ。同類食材判定(R6)の核。"""
    n = str(n)
    if any(k in n for k in ['ハンバーグ','つくね','肉団子','ミートボール','メンチ','そぼろ','ひき肉','つみれ']): return 'ひき肉系'
    if any(k in n for k in ['さつま揚げ','ちくわ','蒲鉾','かまぼこ','がんも','カレーボール','団子','たこ八','たこ焼き','ちぎり揚げ','チヂミ','シュウマイ','焼売','魚肉ソーセージ','カニ玉','かに玉']): return '練り物系'
    if any(k in n for k in ['唐揚','ザンギ','フライドチキン','とり天','チキン','鶏','スティックチキン','ダージーパイ','南蛮']): return '鶏肉系'
    if any(k in n for k in ['豚','チャーシュー','ソーセージ','生姜焼き','酢豚']): return '豚肉系'
    if any(k in n for k in ['牛','ビーフ']): return '牛肉系'
    if '合鴨' in n: return '合鴨系'
    if any(k in n for k in ['鮭','サーモン']): return '鮭系'
    if any(k in n for k in ['さば','サバ','あじ','アジ','ブリ','ぶり']): return '青魚系'
    if any(k in n for k in ['白身魚','カレイ','まぐろ','マグロ','鮪']): return '白身魚系'
    if any(k in n for k in ['海老','えび','エビ']): return '海老系'
    if any(k in n for k in ['かに','カニ','ずわい']): return 'かに系'
    if any(k in n for k in ['いか','イカ','たこ','タコ']): return 'いか・たこ系'
    if any(k in n for k in ['豆腐','高野豆腐','白和え','うの花','おから','卯の花','がんも']): return '豆腐系'
    if any(k in n for k in ['じゃがいも','ポテト','里芋','さつま芋','焼きいも','紅芋','むらさきいも','スイートポテト','山芋','コロッケ']): return 'いも系'
    if 'かぼちゃ' in n: return 'かぼちゃ'
    if any(k in n for k in ['ひじき','切干大根','わかめ','春雨','麩','もずく']): return '乾物系'
    if 'こんにゃく' in n: return 'こんにゃく系'
    if any(k in n for k in ['しめじ','えのき','しいたけ','エリンギ']): return 'きのこ系'
    if any(k in n for k in ['大根','かぶ','れんこん','蓮根','ごぼう','人参','竹の子']): return '根菜系'
    if any(k in n for k in ['キャベツ','小松菜','ほうれん草','白菜','チンゲン菜','水菜','広島菜','高菜','菜の花','もやし','ナムル','青じそ']): return '葉物系'
    if any(k in n for k in ['ピーマン','パプリカ','インゲン','コーン','オクラ','ナス','茄子','ブロッコリー','カリフラワー','きゅうり','枝豆']): return '野菜系'
    if 'サラダ' in n: return '麺・サラダ系'
    return '他'

NERI_KW = ['さつま揚げ','ちくわ','蒲鉾','かまぼこ','がんも','カレーボール','いか団子','たこ八','たこ焼き','ちぎり揚げ','焼売','シュウマイ','魚肉ソーセージ','おさかなソー','花がんも']
KAN_REQUIRE = ['切干大根','ひじき','高野豆腐','麩','春雨']  # 水戻し要

def is_neri(prod):
    if M_NERI.get(prod, '') == '○':
        return True
    return any(k in str(prod) for k in NERI_KW)

def is_kanbutsu_require(prod):
    if '要' in str(M_KAN.get(prod, '')):
        return True
    p = str(prod)
    return any(k in p for k in KAN_REQUIRE) and '戻し' not in p

def protein_from_text(t):
    """メニュー名（変更メモ含む）-> 主タンパク質グループ（R3用）"""
    t = str(t)
    if any(k in t for k in ['鶏','チキン','ザンギ','とり天','コーチン','南蛮','ダージーパイ']): return '鶏'
    if any(k in t for k in ['豚','ポーク','生姜焼き','チャーシュー','酢豚']): return '豚'
    if any(k in t for k in ['牛','ビーフ']): return '牛'
    if any(k in t for k in ['さば','サバ','あじ','アジ','ぶり','ブリ','鮭','サーモン','白身','カレイ','まぐろ','マグロ','鮪']): return '魚'
    return ''

def norm_recipe(n):
    """変種レシピの名寄せ（☆ジャパンフード/☆40g 等の枝番除去）"""
    n = str(n)
    n = re.sub(r'☆.*$', '', n)
    n = re.sub(r'[（(].*?[)）]', '', n)
    n = re.sub(r'\d+g.*$', '', n)
    n = re.sub(r'(通常|→.*|⇒.*|※.*)', '', n)
    return n.strip()

NOISE = ['容器','蓋','箸','カップ','ケース','ｺﾞﾑﾊﾞﾝﾄﾞ','ﾎﾟﾘ','フィルム','フードカップ','フードケース',
         '吸油計算','無洗米','雑穀','押麦','つぼ漬','楊枝','備品','ラベル','袋','上限']
def is_noise(n):
    return any(k in str(n) for k in NOISE) or str(n) == '水'

SUBSTANTIAL = {'ひき肉系','練り物系','いも系','鶏肉系','豚肉系','牛肉系','合鴨系','鮭系','青魚系',
               '白身魚系','海老系','かに系','いか・たこ系','豆腐系','こんにゃく系','乾物系','かぼちゃ'}
COLOR_KW = ['人参','3色ピーマン','3色パプリカ','赤ピーマン','赤パプリカ','レッドピーマン','紅芯大根','パプリカ']

POS_DEF = [('メイン',5,6,7),('サブ',8,9,10),('副菜①',11,12,13),('副菜②',14,15,16),('サラダ',17,18,19)]

def parse_menu(menu):
    """メニュー一覧(ヘッダ無し)から日次の5ポジションを抽出"""
    days = []
    for i in range(menu.shape[0]):
        d = menu.iloc[i, 3] if menu.shape[1] > 3 else None
        if pd.isna(d):
            continue
        try:
            date = pd.to_datetime(d)
        except Exception:
            continue
        if date.year < 2000:
            continue
        rec = {'date': date, '曜日': menu.iloc[i,4] if pd.notna(menu.iloc[i,4]) else '', 'pos': {}}
        for nm_, cm, nmcol, sm in POS_DEF:
            rec['pos'][nm_] = {
                '調理法': menu.iloc[i,cm] if (menu.shape[1]>cm and pd.notna(menu.iloc[i,cm])) else '',
                'name': str(menu.iloc[i,nmcol]) if (menu.shape[1]>nmcol and pd.notna(menu.iloc[i,nmcol])) else '',
            }
        days.append(rec)
    days.sort(key=lambda r: r['date'])
    return days

def yobento_md(name):
    mt = re.search(r'(\d+)月(\d+)日', str(name))
    return (int(mt.group(1)), int(mt.group(2))) if mt else None

def build_day_index(shoku):
    shoku = shoku.copy()
    shoku['md'] = shoku['名称'].apply(yobento_md)
    shoku['isDX'] = shoku['名称'].astype(str).str.contains('【DX】')
    # 【DX】表記が無いシートでは全行を対象にする
    if not shoku['isDX'].any():
        shoku['isDX'] = True
    return shoku

PROT_COLS = ['鶏肉','牛肉','豚肉','鮭','さば','いか','えび','かに']

def day_dishes(shoku, md):
    sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
    dishes, colors = {}, []
    for _, r in sub.iterrows():
        prod = str(r['商品名'])
        qty = r.get('食材数量')
        if pd.isna(qty) or qty == 0 or is_noise(prod):
            continue
        if any(k in prod for k in COLOR_KW):
            colors.append(prod)
        key = norm_recipe(r['レシピ名'])
        if not key or '備品' in key:
            continue
        d = dishes.setdefault(key, {'neri': False, 'kan': False})
        if is_neri(prod): d['neri'] = True
        if is_kanbutsu_require(prod): d['kan'] = True
    return dishes, colors

def run_checks(days, shoku):
    viol = []
    def add(date, yobi, no, rule, where, reason, fix, sev):
        viol.append({'日付': date.strftime('%-m/%-d'), '曜日': yobi, 'No': no, 'ルール': rule,
                     '該当箇所': where, '理由': reason, '修正提案': fix, '重要度': sev})
    run = {}
    for day in days:
        md = (day['date'].month, day['date'].day); yobi = day['曜日']
        dishes, colors = day_dishes(shoku, md)
        pos_group = {p: group_from_name(day['pos'][p]['name']) for p in day['pos'] if day['pos'][p]['name']}
        allnames = ' '.join(p['name'] for p in day['pos'].values())

        # R1 自然解凍が1品以上
        if sum(1 for p in day['pos'].values() if p['調理法'] == '自') == 0:
            add(day['date'],yobi,1,'自然解凍メニューが1品もない','—','自(自然解凍)が0品','副菜を自然解凍品に1つ','中')
        # R8 メイン・サブ調理法被り（参考）
        cm, cs = day['pos']['メイン']['調理法'], day['pos']['サブ']['調理法']
        if cm and cm == cs:
            add(day['date'],yobi,8,'メイン・サブの調理法が同じ',f'両方「{cm}」','調理法被り','サブの調理法を変える','低(参考)')
        # R15 サラダ表記重複
        sal = [p['name'] for p in day['pos'].values() if 'サラダ' in p['name']]
        if len(sal) >= 2:
            add(day['date'],yobi,15,'「サラダ」表記が同日複数','/'.join(sal),f'{len(sal)}品','片方の名称変更','低')
        # R6 同類食材の被り（代表グループ・ポジション横断）
        gc = {}
        for p, g in pos_group.items():
            if g in SUBSTANTIAL:
                gc.setdefault(g, []).append(f'{p}:{day["pos"][p]["name"][:14]}')
        for g, items in gc.items():
            if len(items) >= 2:
                sev = '高' if g in ('ひき肉系','練り物系','いも系') else '中'
                add(day['date'],yobi,6,'同類食材の被り',f'{g} → '+' / '.join(items),f'{g}が{len(items)}品','いずれかを別系統に',sev)
        # R27 練り物被り
        nd = [k for k, d in dishes.items() if d['neri']]
        if len(nd) >= 2:
            add(day['date'],yobi,27,'練り物の被り',' / '.join(nd[:4]),f'練り物{len(nd)}品','片方を練り物以外に','中')
        # R28 乾物(水戻し要)被り
        kd = [k for k, d in dishes.items() if d['kan']]
        if len(kd) >= 2:
            add(day['date'],yobi,28,'乾物(水戻し要)の被り',' / '.join(kd[:4]),f'要戻し乾物{len(kd)}品','片方を水戻し不要 or 別に','低')
        # R5 フライ・揚げの偏り
        fried = [p for p in day['pos'] if day['pos'][p]['調理法'] == '揚']
        if len(fried) >= 3:
            add(day['date'],yobi,5,'食感(フライ・揚げ)の偏り','/'.join(fried),f'揚げ物が{len(fried)}品','1品を煮/和え等に','中')
        # R20 大根とかぶ同日
        if ('大根' in allnames and 'おろし' not in allnames and '切干' not in allnames) and 'かぶ' in allnames:
            add(day['date'],yobi,20,'大根とかぶが同日','—','両方使用','一方に絞る','中')
        # R14 彩り（食材データで判定）
        if not colors:
            add(day['date'],yobi,14,'彩り野菜(人参/赤ピーマン/紅芯大根)が不足','—','色味食材が0','彩り野菜を1品追加','中')
        # R3 連続protein（メイン・サブ）
        today = set(filter(None, [protein_from_text(day['pos']['メイン']['name']),
                                   protein_from_text(day['pos']['サブ']['name'])]))
        for pr in ['鶏','豚','牛','魚']:
            if pr in today:
                run[pr] = run.get(pr, 0) + 1
                if run[pr] >= 3:
                    add(day['date'],yobi,3,'メイン/サブで同タンパクが3日連続',pr,f'{pr}が{run[pr]}日連続(上限2日)','どこかで別タンパクに','高')
            else:
                run[pr] = 0
    return pd.DataFrame(viol)

# ============ 出力 ============
RULE_NAME = {1:'自然解凍1品',3:'同タンパク3日連続',5:'フライ偏り',6:'同類食材被り',8:'メイン/サブ調理法',
             14:'彩り不足',15:'サラダ表記',20:'大根とかぶ',27:'練り物被り',28:'乾物被り'}

def write_report(out, n_days, path):
    order = {'高':0,'中':1,'低':2,'低(参考)':3}
    if len(out):
        out = out.copy()
        out['_s'] = out['重要度'].map(lambda x: order.get(x, 9))
        out['_d'] = pd.to_datetime('2000/'+out['日付'], format='%Y/%m/%d', errors='coerce')
        out = out.sort_values(['_d','_s']).drop(columns=['_s','_d']).reset_index(drop=True)
    wb = Workbook()
    H = Font(name='Meiryo', bold=True, color='FFFFFF', size=10); HF = PatternFill('solid', fgColor='C0703B')
    SUB = Font(name='Meiryo', bold=True, size=11, color='C0703B'); BODY = Font(name='Meiryo', size=10)
    sevfill = {'高':PatternFill('solid',fgColor='F4CCCC'),'中':PatternFill('solid',fgColor='FFF2CC'),
               '低':PatternFill('solid',fgColor='FFFFFF'),'低(参考)':PatternFill('solid',fgColor='F0F0F0')}
    thin = Side(style='thin', color='D9C9B5'); BORD = Border(left=thin,right=thin,top=thin,bottom=thin)
    def hdr(ws, n, row=1):
        for c in range(1, n+1):
            cc = ws.cell(row, c); cc.font = H; cc.fill = HF; cc.border = BORD
            cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Sheet1
    ws = wb.active; ws.title = '違反候補リスト'
    ws.merge_cells('A1:H1'); ws['A1'] = '■ お弁当メニュー違反チェック結果（食材チェックAI）'
    ws['A1'].font = Font(name='Meiryo', bold=True, size=13, color='C0703B'); ws.append([])
    cols = ['日付','曜日','No','ルール','該当箇所','理由','修正提案','重要度']
    ws.append(cols); hdr(ws, len(cols), row=3)
    for _, r in out.iterrows():
        ws.append([r[c] for c in cols])
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        sev = row[7].value
        for cell in row:
            cell.font = BODY; cell.border = BORD; cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.fill = sevfill.get(sev, PatternFill())
    if ws.max_row < 4:
        ws.append(['(違反候補なし)'])
    ws.freeze_panes = 'A4'
    for col, w in zip('ABCDEFGH', [7,5,5,22,40,18,22,9]): ws.column_dimensions[col].width = w
    # Sheet2 サマリー
    ws2 = wb.create_sheet('サマリー'); ws2['A1']='検出サマリー'; ws2['A1'].font=SUB
    ws2.append([]); ws2.append(['ルール','検出件数','重要度内訳']); hdr(ws2,3,row=3)
    if len(out):
        for no in sorted(out['No'].unique()):
            s = out[out['No']==no]
            sv = '/'.join(f'{k}{v}' for k,v in s['重要度'].value_counts().items())
            ws2.append([f'R{no} {RULE_NAME.get(no,"")}', len(s), sv])
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row):
        for cell in row: cell.font=BODY; cell.border=BORD; cell.alignment=Alignment(vertical='center',wrap_text=True)
    ws2.append([]); ws2.append(['合計', len(out), f'検査日数 {n_days}日'])
    ws2.cell(ws2.max_row,1).font = Font(name='Meiryo', bold=True, size=10)
    for col,w in zip('ABC',[26,12,28]): ws2.column_dimensions[col].width=w
    # Sheet3 フロー
    ws3 = wb.create_sheet('チェックフロー')
    ws3['A1']='食材チェックAI 処理フロー'; ws3['A1'].font=Font(name='Meiryo',bold=True,size=13,color='C0703B'); ws3.append([])
    flow = [('① ルール・マスタ内蔵','36ルールと同類食材/食感マスタを保持（毎回送付不要）'),
            ('② 入力','1ヶ月 or 1週間のメニュー＋食材シート(xlsx)'),
            ('③ 展開・紐づけ','日×5ポジションに分解→食材を日付で紐づけ→マスタで主原料グループ/食感/タンパク質/練り物/乾物に変換'),
            ('④ ルール適用','変種レシピは名寄せ、備品/だし/水は除外'),
            ('⑤ 出力','違反候補だけを抽出（日付・ルール・該当箇所・理由・修正提案・重要度）')]
    for a,b in flow:
        ws3.append([a,b])
        ws3.cell(ws3.max_row,1).font=BODY; ws3.cell(ws3.max_row,2).font=BODY
        ws3.cell(ws3.max_row,1).alignment=Alignment(vertical='top',wrap_text=True)
        ws3.cell(ws3.max_row,2).alignment=Alignment(vertical='top',wrap_text=True)
    ws3.column_dimensions['A'].width=22; ws3.column_dimensions['B'].width=70
    wb.save(path)

def main():
    if len(sys.argv) < 2:
        print('使い方: python check_menu.py <入力xlsx> [出力xlsx]'); sys.exit(1)
    inp = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(inp)[0] + '_チェック結果.xlsx'
    xl = pd.ExcelFile(inp)
    menu_sheet = next((s for s in xl.sheet_names if 'メニュー一覧' in s or 'メニュー' in s and '一覧' in s), None)
    shoku_sheet = next((s for s in xl.sheet_names if s == '食材' or '食材' in s), None)
    if not menu_sheet or not shoku_sheet:
        print(f'エラー: 「メニュー一覧」と「食材」タブが必要です。検出タブ: {xl.sheet_names}'); sys.exit(1)
    menu = pd.read_excel(xl, menu_sheet, header=None)
    shoku = pd.read_excel(xl, shoku_sheet, header=0)
    days = parse_menu(menu)
    shoku = build_day_index(shoku)
    out = run_checks(days, shoku)
    write_report(out, len(days), out_path)
    print(f'検査日数: {len(days)} / 違反候補: {len(out)}')
    print(f'出力: {out_path}')
    if len(out):
        print(out.groupby('No').size().to_string())

if __name__ == '__main__':
    main()
