#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
やどかり弁当 メニュー違反チェック（月非依存・汎用版）

アップロードされたワークブックのシート名から月を自動検出し、何ヶ月分でもまとめて処理する。
判定は可能な限りキーワード推測ではなく「実際のマスタデータ（商品ID/レシピID）」で行う方針。

■ 入力（run_all_checks / load_workbook_data の引数）
  xlsx_path         : メインのメニューワークブック（必須）
      - "{月}月使用食材"  : 食材データ（day_csvが無い月のフォールバック用）
      - "{月}月昼夕..."   : 昼夜ペアのメニュー名一覧（No.2/3/5/24/28/29に必要）
      - "{月}月栄養価"    : 日別の栄養価（No.14で優先使用）
  day_csv_paths     : {(月, '昼'|'夜'): csv} 食材CSV。**最重要**。
      商品ID単位で判定する全ルール（No.1/10/12/18/19/20/21/30）の主データ源。
      「N月使用食材」シートが別ライン（DX等）と混ざる問題を避けるため、
      渡された場合は常にこちらを優先する。
  veg_master_path   : 野菜マスタ_テンプレート.xlsx（色列）→ No.9 / No.17
  seasoning_csv_path: 調味料.csv（商品ID）        → No.8 / No.19
  fried_master_path : 食材データ.xlsx             → No.12「調理法（当日揚げ）」シート、
                                                    No.21「禁止食材・調味料該当」シート
  night_csv_paths   : {月: csv} 旧形式の夜食材CSV（day_csv_pathsがあれば不要）

■ 実装済みルール（マスタ照合ベース）
  No.1  同一商品ID（個数カウントの単体商材）をメイン/サブで1週間以内に再使用
  No.3/5 挽肉・鶏豚牛のメイン/サブ同日重複（昼夜別）
  No.4/36 コロッケ連日（同日の複数サイズ登録は対象外）
  No.6/8 1食内の食材/調味料重複（調味料.csv・基礎野菜を除外）
  No.7  大豆系の同一食事内重複（昼夜は半日空きとみなし対象外）
  No.9  野菜マスタの「色」が同じものの2日連続（昼夜別・基礎色/定番食材は除外）
  No.10 単一食材のみの副菜/サラダを1週間空けず再使用
  No.11 自然解凍品が1食に0品
  No.12 当日揚げが3品超（当日揚げレシピIDマスタ照合）
  No.14 栄養素の月平均（「N月栄養価」シート優先。夜はデータ未入手のため昼のみ）
  No.15 健康食材 週1回以上
  No.17 1食で赤・黄・緑を使用（野菜マスタの色列・昼夜別）
  No.18 1食の重量下限（M=212g。容器・カップ重量は含まないため下限側の目安）
  No.19 同じ調味料のみでの味付け禁止
  No.20 だし味付け1品以上
  No.21 禁止食材・調味料（禁止食材マスタ照合／無ければキーワード判定）
  No.22 魚メニュー3日に1回  No.24 白和えの分類
  No.25 かぼちゃ週1回・同曜日4週間  No.26 かにのふわふわ5日以上
  No.27 FD専用魚商材の平日縛り＋★商材の平日夜クオータ（FDメニュールール準拠）
  No.28 本日の魚料理は平日夜  No.29 おまかせ月2回以上
  No.30 野菜の使用間隔（FDメニュールール（野菜）シート準拠・商品ID単位。
        メニュー名記載時は必要日数2倍、芋類/かぼちゃは昼夜連続OK）
  No.31 マッシュ系同日重複

■ 未対応
  No.2  実装済みだがユーザー指示によりALL_RULESから除外中
  No.23 実装済みだがユーザー指示によりチェック対象外（ALL_RULESから除外中）
  No.13 盛付工程（工程データが無いためスキップ）
  No.16 固形2種まで（判定基準となるマスタが未確定）
  No.32〜35, 37（未着手）

■ NG時の代替え案
  違反行の「修正提案」列には、可能な場合『直近使われていない具体的な代替商材/レシピ名』を出す。
  同系統(cm.group_from_name)・同色・同カテゴリ等の候補から、その日時点で最も長く
  使われていないものを選ぶ（終売商品は候補から除外）。
"""
import os
import re
import sys
import json
import datetime
from collections import Counter
import pandas as pd
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, 'scripts'))
import check_menu as cm  # noqa: E402

# ---- check_menu.group_from_name のパッチ（同梱スクリプトは読み取り専用のため、こちら側で拡張） ----
# 実データ（新構成ルール確認.xlsx 7-8月）を全件確認して見つかったキーワード漏れを追加。
# 見つかり次第ここに追記していく。
_ORIG_GROUP_FROM_NAME = cm.group_from_name
_GROUP_PATCH_HIKINIKU = ['肉詰め']
_GROUP_PATCH_SHIROMI = ['タラ', 'たら', '鯛', 'サワラ', 'さわら', '金目鯛', 'ほっけ', 'ホッケ']


def _patched_group_from_name(n):
    n = str(n)
    if any(k in n for k in _GROUP_PATCH_HIKINIKU):
        return 'ひき肉系'
    if any(k in n for k in _GROUP_PATCH_SHIROMI):
        return '白身魚系'
    return _ORIG_GROUP_FROM_NAME(n)


cm.group_from_name = _patched_group_from_name

WD_JP = ['月', '火', '水', '木', '金', '土', '日']
# 各ルール関数が返す内部形式（曜日欄には slot や 曜日 が混在する）
cols_std = ['日付', '曜日', 'No', 'ルール', '該当箇所', '理由', '修正提案', '重要度']
# 最終出力の形式。「曜日」ではなく「昼夜」を出す（ユーザー指定）
OUT_COLS = ['日付', '昼夜', 'No', 'ルール', '該当箇所', '理由', '修正提案', '重要度']


def _derive_slot(row):
    """行の内容から昼/夜を判定する。
    ルール関数によって slot の持ち方が違う（曜日欄に'昼'/'夜/火'、該当箇所に'[昼]'/'夜サブ:'等）ため、
    曜日欄→該当箇所→理由 の順に探す。昼夜を区別しない日単位ルールは'昼夜'、
    月単位の集計行は'-'を返す。"""
    wd = str(row.get('曜日', ''))
    if '昼' in wd and '夜' in wd:
        return '昼夜'
    if '昼' in wd:
        return '昼'
    if '夜' in wd:
        return '夜'
    # 該当箇所は「夜サブ:○○ ← 前回 7/2 昼サブ:△△」のように前回分の昼夜も含むため、
    # 最初に出てくる方（＝その違反行の当該食事）を採用する。
    for col in ('該当箇所', '理由'):
        s = str(row.get(col, ''))
        i_l, i_n = s.find('昼'), s.find('夜')
        if i_l < 0 and i_n < 0:
            continue
        if i_n < 0 or (0 <= i_l < i_n):
            return '昼'
        return '夜'
    return '-' if str(row.get('日付', '')).endswith('月') else '昼夜'


def _date_sort_key(v):
    """'7/1'→(7,1,1) / '7月(月次)'→(7,0,0)（月次集計行はその月の先頭）で並べ替えるキーを返す。"""
    s = str(v)
    m = re.match(r'^(\d{1,2})月', s)
    if m:
        return (int(m.group(1)), 0, 0)
    m = re.match(r'^(\d{1,2})/(\d{1,2})', s)
    if m:
        return (int(m.group(1)), int(m.group(2)), 1)
    return (99, 99, 9)

BASE_SEASONING_KW = ['だし', '醤油', '塩こしょう', '片栗粉', '上白糖', 'コンソメ', 'ウスターソース',
                      'ごま油', '味覇', 'シャンタン', 'みりん', '料理酒', '食塩', 'こしょう', '胡椒',
                      '酢', 'マヨネーズ', 'ケチャップ', 'ソース', 'たれ', 'ダレ', 'あん', 'スパイス',
                      '昆布茶', 'スープ', '味噌', '砂糖', 'サラダ油', '唐辛子', 'わさび',
                      '辛子', 'からし', 'ラー油', '豆板醤', 'オイスター', 'ポン酢', 'カレー粉', '小麦粉']
# 「下ごしらえ済みの基礎野菜」（No.6 同一食材複数レシピ重複の判定から除外する）
# ユーザー確認済み：ネギ・生姜・人参・玉ねぎ類のような、どの料理にも使う土台の野菜は
# 調味料と同様に重複してもNo.6の違反対象にしない
BASE_VEG_KW = ['白ネギ', 'ネギ', 'ねぎ', '玉ねぎ', 'たまねぎ', 'オニオン', '生姜', 'しょうが', 'ショウガ',
               '人参', 'にんじん', 'えのき', 'インゲン', 'いんげん', 'ピーマン', 'パプリカ', 'にんにく', 'ニンニク']
SOY_KW = ['豆腐', 'がんも', 'おから', '卯の花', 'うの花', '豆乳', '高野豆腐', '厚揚げ', '油揚げ', '生揚げ', '湯葉', '大豆', '納豆']
FISH_KW = ['鮭', 'サーモン', 'さば', 'サバ', 'あじ', 'アジ', 'ぶり', 'ブリ', '白身魚', 'カレイ', 'かれい',
           'まぐろ', 'マグロ', '鮪', 'さわら', 'サワラ', 'たら', 'タラ', 'いわし', 'イワシ', 'さんま', 'サンマ',
           'ほっけ', 'ホッケ', '金目鯛', '鯛', 'タイ', 'アブラカレイ', 'ししゃも', 'シシャモ']
NG_WORDS = ['餅', 'ヤングコーン', 'パスタこんにゃく', '旨辛ジャン', '麻婆']
HEALTH_KW = ['アジ', 'あじ', 'イワシ', 'いわし', 'サバ', 'さば', 'サンマ', 'さんま', 'マグロ', 'まぐろ',
             'レモン', '酢の物', 'ほうれん草', 'あさり', 'アサリ', '豆乳', 'ひじき']
EAT_NG = {
    'イカ': '硬さ(歯で嚙み切れない)', 'ビーフジャーキー': '硬さ(歯で嚙み切れない)',
    'こんにゃく': '弾力が強すぎる', 'ナタデココ': '弾力が強すぎる', 'ホルモン': '弾力が強すぎる',
    'ごぼう': '繊維質', 'セロリ': '繊維質', '牛すじ': '繊維質',
    'もち': '口に張り付く', '海苔': '口に張り付く', 'ウエハース': '口に張り付く',
    '鶏むね': 'パサつく', '焼き魚': 'パサつく', '茹でたまご': 'パサつく', 'ゆで卵': 'パサつく',
    'おから': 'ぼそぼそする', '玄米': 'ぼそぼそする', 'スコーン': 'ぼそぼそする',
    '春巻き': '口腔内を傷つける', 'チップス': '口腔内を傷つける', '乾燥した小魚': '口腔内を傷つける', '田作り': '口腔内を傷つける',
}
LOOKALIKE_PAIRS = [
    (['海老カツ', '海老タラカツ'], 8),
    (['塩じゃがコロッケ', 'ビーフ入りコロッケ'], 8),
    (['さつまいもコロッケ', 'かぼちゃコロッケ'], 8),
    (['ガツンとジューシーメンチ', 'ビーフ入りメンチカツ'], 8),
]
# 「枠タイトル＋☆(または※)＋実際の料理名」形式のメニュー。
# 枠タイトルだけを見ると別料理が同一判定される誤検知の元になるため、
# canon_name ではこれらのプレフィックスを検出したら☆/※より後ろの実料理名を名寄せキーにする。
# (新構成ルール確認.xlsx 7月/8月の実データを全件走査して確認した一覧。新しい枠タイトルが
#  出てきた場合はここに追加する)
FRAME_TITLES = ['お楽しみの1品', 'お楽しみの揚げ物', '店主おすすめの1品', 'おまかせの副菜', '本日の魚料理']
VEG_TIERS = [
    (1, ['ほうれん草']),
    (1, ['インゲンカット', 'ミニミニブロッコリー', 'オクラ', 'ささがきごぼう', '大根乱切り', '冷凍かぶ', '竹の子千切り', 'スナップピース']),
]

# No.30用: 「FDメニュールール（野菜）」シート
# (https://docs.google.com/spreadsheets/d/1w6ck7gAUbJIOOlDODM58QKj6nkBc2WSX5T0_Cpv7QBY/edit?gid=1671677088)
# の内容をもとにした間隔マスタ（ユーザー確認済み・2026/8/5時点の内容を反映）。
# 「昼･夜、夜･昼使用可能」「半日空いていればOK」の2階層（ほうれん草を除く）は、日付単位の
# チェックでは実質常に満たされるため対象外とし、1.5日以上の間隔が必要な階層のみ判定する。
# 各要素: (match_type('id'/'name'), key(商品ID or レシピ名に含むキーワード), 名寄せ用キーワード,
#          基本必要日数, メニュー名に記載がある場合の必要日数, 昼夜連続(半日)は例外的にOKか, 表示名)
# 同シートの「昼･夜、夜･昼使用可能」「連続OK（半日空いていればOK）」階層の商品ID。
# 間隔制約がほぼ無いため、No.30違反時の『代わりに使える野菜』候補として使う。
VEG_FLEXIBLE_IDS = [
    3001891, 3002435, 3001871, 3002300, 3002403, 3001844, 3002158,     # 昼夜/夜昼 使用可能
    3001945, 3001944, 3002217, 3000080, 3000914, 3000001, 3002357,     # 連続OK（半日）
    3002248, 3003055, 3002192, 3003057, 3001890, 3000024, 3003009, 3001403,
]

VEG_TIER_MASTER = [
    # 半日階層だが、ほうれん草のみ1.5日に格上げ（シートD11注記）
    ('id', 3002349, 'ほうれん草', 1.5, 3, False, '自然解凍 ほうれん草IQF 500g'),
    # 1.5日空けばOK階層
    ('id', 3002303, 'インゲンカット', 1.5, 3, False, 'インゲンカット(要加熱) 500g'),
    ('id', 3003003, 'ミニミニブロッコリー', 1.5, 3, False, '冷凍ミニミニブロッコリー 500g'),
    ('id', 3002371, 'オクラ', 1.5, 3, False, 'オクラスライス 500g'),
    ('id', 3002267, 'ささがきごぼう', 1.5, 3, False, '前川 冷凍ささがきごぼう 500g'),
    ('id', 3002290, '大根乱切り', 1.5, 3, False, '大根乱切り 500g'),
    ('id', 3002287, 'かぶ', 1.5, 3, False, '冷凍かぶ 銀杏切り 500g'),
    ('id', 3001594, '竹の子', 1.5, 3, False, '冷凍竹の子千切り 500g'),
    ('id', 3000084, 'スナップ', 1.5, 3, False, 'アイガースナップピース 500g'),
    # 3日空けばOK（メニュー名に無ければ2日でも可）階層。芋(じゃがいも/さつまいも/里芋)とかぼちゃは
    # 昼夜連続使用（半日空き）は例外的にOK（シートD42注記）。れんこん・油調ナスは対象外。
    ('id', 3002377, '油調ナス', 2, 3, False, '油調ナス(自然解凍) 500g'),
    ('id', 3003062, ('さといも', '里芋'), 2, 3, True, 'さといもSS 500g'),
    ('id', 3002214, 'じゃがいも', 2, 3, True, '乱切りじゃがいも 500g'),
    ('id', 3001814, ('さつま芋', 'さつまいも'), 2, 3, True, 'ひとくち焼きいも(さつま芋)'),
    ('id', 3001964, 'スイートポテト', 2, 3, True, 'スイートポテト 1kg'),
    ('id', 3002355, 'れんこん', 2, 3, False, 'れんこん乱切り 500g'),
    # 2日空けばOK（かぼちゃ3商品）階層。同上、昼夜連続はOK。
    ('name', 'かぼちゃ', 'かぼちゃ', 2, 2, True, 'かぼちゃ系（皮つき乱切り/煮物/栗南瓜コロッケ等）'),
]
# No.27用: 「FDメニュールール」シート（★マーク商品のうち備考欄に「平日夜に◯回は入れる」という
# 明示クオータがある商品のみ・ユーザー確認済みスコープ）。
# (https://docs.google.com/spreadsheets/d/1w6ck7gAUbJIOOlDODM58QKj6nkBc2WSX5T0_Cpv7QBY/edit?gid=1597935310)
# 各要素: (商品ID or None, 名寄せキーワード, 月内の平日夜 最低使用回数, 枠)
FD_WEEKDAY_NIGHT_QUOTA = [
    (3002318, '7品目具材の豆腐ハンバーグ', 2, 'サブ'),
    (3001677, 'ピーマン肉詰めフライ', 1, 'メイン'),
    (3002409, '三元豚ロｰストンカツ', 2, 'メイン'),
    (3001155, 'チキン八幡巻', 2, 'メイン'),
    (None, '魚弁当', 1, 'サブ'),
    (3001449, 'かにのふわふわ豆腐', 1, 'サブ'),
    (3001909, '生姜入り豆腐ステーキ', 2, 'サブ'),
    (3002155, '豆乳と野菜のふわふわ真丈', 2, '副菜'),
    (3003053, '栗かぼちゃ旨煮', 3, '副菜'),
]

FISH_FD_ONLY = ['いわしの梅煮', 'マスの塩焼き', 'サーモン塩焼き', 'ぶりのみぞれ', 'さばの味噌煮',
                'さわらの西京焼き', 'タラの香草焼き', 'あじみりん焼き', 'あじの塩焼き']
NUTRI_BOUNDS = {'昼': {'kcal': (415, 455), 'salt_max': 3.8, 'protein_min': 12},
                '夜': {'kcal': (245, 275), 'salt_max': 3.0, 'protein_min': 12}}
# No.17: 1食につき赤・黄・緑を必ず使用（キーワード方式・暫定版。商材マスタの「色」列が整備され次第、
# そちらを正とする判定に切り替える想定）
RED_KW = ['人参', 'にんじん', '3色ピーマン', '3色パプリカ', '赤ピーマン', '赤パプリカ', 'レッドピーマン',
          '紅芯大根', 'トマト', 'ミニトマト', '紅生姜', 'いちご']
YELLOW_KW = ['3色ピーマン', '3色パプリカ', '黄パプリカ', '黄ピーマン', 'イエローピーマン', 'コーン', 'とうもろこし',
             'かぼちゃ', '卵', 'たまご', '玉子', 'たくあん', 'パイナップル', 'レモン']
GREEN_KW = ['ほうれん草', '小松菜', 'ブロッコリー', 'いんげん', 'インゲン', 'オクラ', 'きゅうり', 'キュウリ',
            'ピーマン', '枝豆', 'さやえんどう', 'スナップ', 'アスパラ', '春菊', '水菜', 'チンゲン菜']


def _nfkc(s):
    """半角カタカナ(ｺﾝｿﾒ等)・全角英数字などを正規化する。商品名データには半角カタカナ表記が
    多く混ざっており、全角キーワードでのin判定が素通りしてしまう不具合があったため、
    キーワード一致判定の前段に必ずこれを通す。"""
    import unicodedata
    return unicodedata.normalize('NFKC', str(s))


def _normalize_text_columns(df, cols=('商品名', 'レシピ名', '名称')):
    """食材データの商品名/レシピ名/名称列を読み込み直後にNFKC正規化する（半角カタカナ対策）。
    これにより、以降のキーワード一致判定（is_base_seasoning等）が全角/半角の表記ゆれを
    気にせず動くようになる。データソースの入口1箇所で正規化することで、個々のルール
    関数側での対応漏れを防ぐ。"""
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = df[c].astype(str).map(_nfkc)
    return df


def is_base_seasoning(prod):
    return any(k in _nfkc(prod) for k in BASE_SEASONING_KW)


def is_base_veg(prod):
    return any(k in _nfkc(prod) for k in BASE_VEG_KW)


def is_soy(name):
    return any(k in _nfkc(name) for k in SOY_KW)


def is_fish(name):
    return any(k in _nfkc(name) for k in FISH_KW)


def is_health(name):
    return any(k in _nfkc(name) for k in HEALTH_KW)


def _to_date(v):
    """セルの値をdatetimeに変換する。日付書式のセルはdatetime.datetimeで返るが、
    書式が数値のままの一部ファイルではExcelのシリアル値（整数/浮動小数）で返ってくるため、
    その場合も1899-12-30起点で変換する。"""
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool) and v > 0:
        try:
            return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=v)
        except (OverflowError, ValueError):
            return None
    return None


def canon_name(name):
    name = str(name)
    stripped = name.replace('　', ' ').strip()
    # 枠タイトル形式（お楽しみの1品☆xxx 等）は、枠タイトルではなく☆/※の後ろの
    # 実料理名を名寄せキーにする（枠タイトルだけで名寄せすると別料理を同一視してしまうため）
    for ft in FRAME_TITLES:
        if stripped.startswith(ft):
            rest = stripped[len(ft):]
            rest = re.sub(r'^[\s☆※]+', '', rest)
            if rest:
                for group, _ in LOOKALIKE_PAIRS:
                    for g in group:
                        if g in rest:
                            return group[0]
                return cm.norm_recipe(rest) or rest
            break
    for group, _ in LOOKALIKE_PAIRS:
        for g in group:
            if g in name:
                return group[0]
    return cm.norm_recipe(name)


AI_MODEL = 'claude-haiku-4-5-20251001'
_AI_SIM_CACHE = {}


def get_anthropic_client():
    """Streamlit CloudのSecrets（ANTHROPIC_API_KEY）か環境変数からAPIキーを取得してクライアントを作る。
    キーが無い/anthropicパッケージが未インストールならNoneを返す（呼び出し側はAI判定をスキップする）。"""
    try:
        import anthropic
    except ImportError:
        return None
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        try:
            import streamlit as st
            api_key = st.secrets.get('ANTHROPIC_API_KEY')
        except Exception:
            api_key = None
    if not api_key:
        return None
    try:
        return anthropic.Anthropic(api_key=api_key)
    except Exception:
        return None


def ai_similar_candidates(client, new_name, candidate_names, model=AI_MODEL):
    """new_name と candidate_names の中から『同じ/酷似した料理』とAIに判定させ、
    酷似すると判定された候補名の集合を返す。API呼び出しに失敗した場合は空集合（＝AI判定なし）。
    同じ組み合わせは再度APIを呼ばずキャッシュから返す。"""
    candidate_names = [c for c in dict.fromkeys(candidate_names) if c and c != new_name]
    if not candidate_names or client is None:
        return set()
    cache_key = (new_name, tuple(sorted(candidate_names)))
    if cache_key in _AI_SIM_CACHE:
        return _AI_SIM_CACHE[cache_key]
    numbered = '\n'.join(f'{i + 1}. {n}' for i, n in enumerate(candidate_names))
    prompt = (
        'あなたは高齢者向け配食弁当のメニュー構成をチェックしている担当者です。\n'
        '以下の「新しく使う商品名」が、「直近使用済みリスト」の中の商品と、'
        '同じ料理・酷似した料理（主原料・調理法・味付けの系統が同じ）とみなせる場合、'
        'その番号をすべて挙げてください。\n'
        '例：「コク深い味噌の甘辛ハンバーグ」と「おろしハンバーグ」は、どちらも'
        '「ハンバーグ」が主役なので酷似とみなします。「大葉おろしチキンカツ」と'
        '「おろしチキンカツ」も、大葉の有無だけの違いなので酷似とみなします。\n'
        '単に同じ食材カテゴリ（両方とも肉料理、等）というだけでは酷似とみなさず、'
        '料理として同じ系統（ハンバーグ同士、コロッケ同士、唐揚げ同士 等）の場合のみ酷似と判定してください。\n\n'
        f'新しく使う商品名: {new_name}\n\n'
        f'直近使用済みリスト:\n{numbered}\n\n'
        '酷似すると判定した番号のみをJSON配列で出力してください（例: [1, 3]）。'
        '酷似するものが無ければ [] と出力してください。JSON以外は一切出力しないでください。'
    )
    try:
        resp = client.messages.create(
            model=model, max_tokens=200,
            messages=[{'role': 'user', 'content': prompt}],
        )
        text = resp.content[0].text.strip()
        m = re.search(r'\[.*?\]', text, re.S)
        idxs = json.loads(m.group(0)) if m else []
        result = {candidate_names[i - 1] for i in idxs if isinstance(i, int) and 1 <= i <= len(candidate_names)}
    except Exception:
        result = set()
    _AI_SIM_CACHE[cache_key] = result
    return result


class MenuData:
    """1つのアップロードから抽出した、月ごとのデータをまとめて保持する"""

    def __init__(self):
        self.months = []                 # [7, 8, ...]
        self.shoku = {}                  # month -> DataFrame (昼, build_day_index済み)
        self.shoku_night = {}            # month -> DataFrame (夜, build_day_index済み) ※無ければキー無し
        self.menu_lunch_tagged = {}      # month -> (df, pos_def) 調理法/タグ付き昼メニュー ※無ければキー無し
        self.rows = []                   # [(date, weekday_jp, slot, pos, name), ...] 全月分
        self.warnings = []               # 注記（画面に表示する）
        self.date_range = None
        self.ai_client = None            # Anthropic clientが設定されていればAI酷似判定を使う（予備・現状未使用）
        self.day_csv = {}                # {(month, '昼'/'夜'): DataFrame}　'md'列付き。No.1(商品ID単位)判定用
        self.nutrition_shoku = {}        # month -> DataFrame（カロリー等の列を持つ「N月使用食材」シートがあれば。No.14専用）
        self.nutrition_daily = {}        # month -> DataFrame(date,kcal,protein,salt)。「N月栄養価」シート由来。No.14専用（優先使用）
        self.veg_color_map = []          # [(正規化名, {色,...}), ...] 野菜マスタ由来。No.9専用
        self.seasoning_ids = set()       # 調味料.csv由来の商品ID集合。No.8専用（キーワードでなく実データで判定）
        self.fried_recipe_ids = set()    # 食材データ.xlsx「調理法（当日揚げ）」由来のレシピID集合。No.12専用
        self.seasoning_names = {}        # 調味料.csv由来の 商品ID -> 商品名。No.19の代替案提案用
        self.ng_product_ids = set()      # 食材データ.xlsx「禁止食材・調味料該当」由来の商品ID集合。No.21専用
        self.ng_product_names = {}       # 同シート由来の 商品ID -> 商品名
        self.recipe_nutrition = None     # 食材データ.xlsx由来のレシピ別栄養価。No.14の代替メニュー提案用
        self._usage_hist = None          # 商品名(NFKC) -> 使用日リスト（キャッシュ、代替案提案用）


def _find_month_sheets(sheet_names, suffix_regex):
    result = {}
    for sn in sheet_names:
        m = re.match(r'^(\d{1,2})月' + suffix_regex, sn)
        if m:
            month = int(m.group(1))
            result.setdefault(month, sn)
    return result


def _detect_lunchdinner_layout(ws, date_col=2, label_col=1, sample_rows=30):
    """列構成が「タグ,名前」のペア型(7月昼夕比較: 名前列=4,6,8,10,12、5品分で列Lまで使う)か、
    「名前のみ」の単列型(8月昼夕: 名前列=4,5,6,7,8、5品分で列Hまでしか使わない)かを判定する。
    ペア型は5品目の名前が列12(L)付近まで伸びるのに対し、単列型は列8(H)までしか埋まらない
    ため、列9〜13(I〜M)に値があるかどうかを判定材料にする（奇偶列だけを見ると
    単列型でも偶然半分が奇数列に当たってしまい誤判定するため使わない）。"""
    cur_date = None
    paired_votes = 0
    straight_votes = 0
    checked = 0
    for r in range(1, ws.max_row + 1):
        if checked >= sample_rows:
            break
        label = ws.cell(row=r, column=label_col).value
        if label is None:
            continue
        d = ws.cell(row=r, column=date_col).value
        dd = _to_date(d)
        if dd is not None:
            cur_date = dd
        if cur_date is None:
            continue
        vals_3_8 = [ws.cell(row=r, column=c).value for c in range(3, 9)]
        if sum(1 for v in vals_3_8 if v not in (None, '')) == 0:
            continue
        checked += 1
        beyond8 = any(ws.cell(row=r, column=c).value not in (None, '') for c in range(9, 14))
        if beyond8:
            paired_votes += 1
        else:
            straight_votes += 1
    if paired_votes >= straight_votes:
        return {4: 'メイン', 6: 'サブ', 8: '副菜1', 10: '副菜2', 12: 'サラダ'}
    return {4: 'メイン', 5: 'サブ', 6: '副菜1', 7: '副菜2', 8: 'サラダ'}


def _parse_lunch_dinner_sheet(ws, date_col=2, label_col=1):
    """「N月昼夕...」系シートを (date, weekday_jp, slot, pos, name) のリストにする。
    列構成は 7月昼夕比較(タグ,名前 の2列セット×5) と 8月昼夕(名前のみ×5) の両パターンに対応。"""
    pos_map = _detect_lunchdinner_layout(ws, date_col, label_col)
    rows = []
    cur_date = None
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if label is None:
            continue
        d = ws.cell(row=r, column=date_col).value
        dd = _to_date(d)
        if dd is not None:
            cur_date = dd
        if cur_date is None:
            continue
        slot = '夜' if '夜' in str(label) else ('昼' if '昼' in str(label) else str(label))
        for c, pos in pos_map.items():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                rows.append((cur_date.date(), WD_JP[cur_date.weekday()], slot, pos, _nfkc(v.strip())))
    return rows


def load_veg_color_map(veg_master_path):
    """野菜マスタ_テンプレート.xlsx（列の説明・対応ルールシート準拠）を読み込み、
    [(正規化名, {色1, 色2, ...}), ...] のリストを返す（色が「赤・黄・緑」等の場合は分割）。
    正規化名が長い順に並べ替えて返す（後段の部分一致で、短い名前が長い名前の一部に
    誤って先にマッチしないようにするため。例：「ピーマン」が「ピーマン肉詰めフライ」に
    誤爆しないよう、より具体的な名前を優先する）。"""
    wb = openpyxl.load_workbook(veg_master_path, data_only=True)
    sheet = next((s for s in wb.sheetnames if 'マスタ' in s and 'テンプレート' in s), wb.sheetnames[0])
    ws = wb[sheet]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        name = row[2] if len(row) > 2 else None
        color = row[3] if len(row) > 3 else None
        if not name or not color:
            continue
        colors = set(re.split(r'[・,、]', str(color)))
        out.append((_nfkc(str(name)), colors))
    out.sort(key=lambda x: -len(x[0]))
    return out


def veg_colors_for(name, veg_color_map):
    """商品名/レシピ名からマッチする野菜マスタの色集合を返す（複数ヒットしうる）"""
    n = _nfkc(str(name))
    colors = set()
    for veg_name, cset in veg_color_map:
        if veg_name in n:
            colors |= cset
    return colors


def load_seasoning_ids(seasoning_csv_path):
    """調味料.csv（商品ID/商品名/カテゴリ名列を含む）から商品ID集合を返す。No.8専用。
    キーワード一致(is_base_seasoning)より正確に『実際に調味料マスタに登録された商品』を判定できる。"""
    df = pd.read_csv(seasoning_csv_path)
    return set(pd.to_numeric(df['商品ID'], errors='coerce').dropna().astype(int))


def load_seasoning_names(seasoning_csv_path):
    """調味料.csv から 商品ID -> 商品名 の辞書を返す。No.19の代替案提案用。"""
    df = pd.read_csv(seasoning_csv_path)
    df = _normalize_text_columns(df, cols=('商品名',))
    ids = pd.to_numeric(df['商品ID'], errors='coerce')
    out = {}
    for i, name in zip(ids, df['商品名']):
        if pd.isna(i):
            continue
        out[int(i)] = str(name)
    return out


def _find_nutrition_daily_sheet(wb_raw, sheet_name):
    """「N月栄養価」シートから日別のエネルギー/たんぱく質/食塩相当量を抽出する。
    ヘッダ位置（7月栄養価はB1に説明文＋列名なしのタイトル行、8月栄養価は明示的な列名行）が
    月によって微妙に違うため、『エネルギー』という文字列を含むセルを探してその行を
    ヘッダ行とみなし、列位置を動的に検出する（ユーザー指定：7月栄養価/8月栄養価シート参照）。"""
    ws = wb_raw[sheet_name]
    header_row, col_kcal, col_protein, col_salt = None, None, None, None
    for r in range(1, min(ws.max_row, 5) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and 'エネルギー' in v:
                header_row, col_kcal = r, c
            if isinstance(v, str) and 'たんぱく質' in v:
                col_protein = c
            if isinstance(v, str) and '食塩' in v:
                col_salt = c
        if header_row:
            break
    if not header_row or not col_kcal:
        return pd.DataFrame(columns=['date', 'kcal', 'protein', 'salt'])
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        dd = _to_date(d)
        if dd is None:
            continue
        kcal = ws.cell(row=r, column=col_kcal).value
        protein = ws.cell(row=r, column=col_protein).value if col_protein else None
        salt = ws.cell(row=r, column=col_salt).value if col_salt else None
        if kcal is None:
            continue
        rows.append({'date': dd.date(), 'kcal': kcal, 'protein': protein, 'salt': salt})
    return pd.DataFrame(rows)


def load_fried_recipe_ids(shoku_data_path, sheet='調理法（当日揚げ）'):
    """食材データ.xlsx の「調理法（当日揚げ）」シート（ヘッダはB1に説明文、B3行目が列名、
    A4行目以降が実データ、A列=レシピID）からレシピID集合を返す。No.12専用。"""
    wb = openpyxl.load_workbook(shoku_data_path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df = df.dropna(subset=['レシピID'])
    return set(df['レシピID'].astype(int))


def load_ng_product_ids(shoku_data_path, sheet='禁止食材・調味料該当'):
    """食材データ.xlsx の「禁止食材・調味料該当」シート（ヘッダはB1に説明文、B3行目が列名
    '食材ID'/'商品名'、A4行目以降が実データ）から (商品ID集合, {商品ID: 商品名}) を返す。No.21専用。
    従来のキーワード判定(NG_WORDS)より正確に『実際に禁止マスタに登録された商品』を判定できる。"""
    wb = openpyxl.load_workbook(shoku_data_path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df = df.dropna(subset=['食材ID'])
    ids = pd.to_numeric(df['食材ID'], errors='coerce').dropna().astype(int)
    names = df.get('商品名')
    name_map = {}
    if names is not None:
        for i, nm in zip(ids, names.loc[ids.index]):
            name_map[int(i)] = str(nm)
    return set(ids), name_map


def load_recipe_nutrition(shoku_data_path):
    """食材データ.xlsx の「レシピID/名称/カロリー/たんぱく質/食塩相当量」列を持つ全シート
    （調理法（当日揚げ）／調理法（前日揚げ）／マッシュ該当 等）を統合し、
    レシピ単位の栄養価一覧DataFrameを返す。No.14の具体的な代替メニュー提案に使う。"""
    wb = openpyxl.load_workbook(shoku_data_path, data_only=True)
    frames = []
    for s in wb.sheetnames:
        ws = wb[s]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        if not rows or not rows[0]:
            continue
        cols = [str(c) if c is not None else '' for c in rows[0]]
        if 'レシピID' not in cols or 'カロリー' not in cols:
            continue
        df = pd.DataFrame(rows[1:], columns=cols)
        keep = [c for c in ['レシピID', '名称', 'カロリー', 'たんぱく質', '食塩相当量'] if c in df.columns]
        df = df[keep].dropna(subset=['レシピID', '名称'])
        frames.append(df)
    if not frames:
        return pd.DataFrame(columns=['レシピID', '名称', 'カロリー', 'たんぱく質', '食塩相当量'])
    out = pd.concat(frames, ignore_index=True).drop_duplicates(subset='レシピID')
    for c in ['カロリー', 'たんぱく質', '食塩相当量']:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors='coerce')
        else:
            out[c] = pd.NA
    out['名称'] = out['名称'].astype(str).apply(_nfkc)
    return out


def load_workbook_data(xlsx_path, night_csv_paths=None, day_csv_paths=None, veg_master_path=None,
                        seasoning_csv_path=None, fried_master_path=None):
    """xlsx_path: メインのメニューワークブック。
    night_csv_paths: {month: csv_path} 夜食材CSV（任意）。
    day_csv_paths: {(month, '昼'|'夜'): csv_path} 商品ID単位のNo.1判定用CSV（任意）。
        「N月使用食材」シートが別商品ライン（例:DXライン）と混ざっている場合など、
        メニュー一覧と同じラインの食材CSVを明示的に渡したい時に使う。
    veg_master_path: 野菜マスタ_テンプレート.xlsx のパス（任意）。No.9（見た目色の2日連続判定）用。
    seasoning_csv_path: 調味料.csv のパス（任意）。No.8（食材+調味料の重複判定）で、
        キーワードでなく実際の調味料マスタの商品IDで調味料かどうかを判定するために使う。"""
    night_csv_paths = night_csv_paths or {}
    day_csv_paths = day_csv_paths or {}
    data = MenuData()
    if veg_master_path:
        try:
            data.veg_color_map = load_veg_color_map(veg_master_path)
        except Exception as e:
            data.warnings.append(f'野菜マスタの読み込みに失敗しました（{e}）')
    if seasoning_csv_path:
        try:
            data.seasoning_ids = load_seasoning_ids(seasoning_csv_path)
            data.seasoning_names = load_seasoning_names(seasoning_csv_path)
        except Exception as e:
            data.warnings.append(f'調味料.csvの読み込みに失敗しました（{e}）')
    if fried_master_path:
        try:
            data.fried_recipe_ids = load_fried_recipe_ids(fried_master_path)
        except Exception as e:
            data.warnings.append(f'食材データ.xlsx（当日揚げ）の読み込みに失敗しました（{e}）')
        try:
            data.ng_product_ids, data.ng_product_names = load_ng_product_ids(fried_master_path)
        except Exception as e:
            data.warnings.append(f'食材データ.xlsx（禁止食材・調味料該当）の読み込みに失敗しました（{e}）')
        try:
            data.recipe_nutrition = load_recipe_nutrition(fried_master_path)
        except Exception as e:
            data.warnings.append(f'食材データ.xlsx（レシピ別栄養価）の読み込みに失敗しました（{e}）')
    for key, path in day_csv_paths.items():
        try:
            df = pd.read_csv(path)
            df = _normalize_text_columns(df)
            df['md'] = df['名称'].apply(cm.yobento_md)
            data.day_csv[key] = df
        except Exception as e:
            data.warnings.append(f'{key}：day_csv読み込みに失敗しました（{e}）')
    xl = pd.ExcelFile(xlsx_path)
    wb_raw = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = xl.sheet_names
    shoku_sheets = _find_month_sheets(sheet_names, r'使用食材$')
    lunchdinner_sheets = _find_month_sheets(sheet_names, r'昼夕')
    lunch_tagged_sheets = _find_month_sheets(sheet_names, r'昼$')
    night_sheets = _find_month_sheets(sheet_names, r'.*夜食材$')
    nutrition_daily_sheets = _find_month_sheets(sheet_names, r'栄養価$')
    months = sorted(set(shoku_sheets) | set(lunchdinner_sheets))
    if not months:
        raise ValueError('「N月使用食材」「N月昼夕...」形式のシートが見つかりませんでした。シート名をご確認ください。')
    data.months = months
    NUTRI_COLS = ['カロリー', 'たんぱく質', '食塩相当量']
    for month in months:
        # 食材ベースの判定（No.4/6/7/21/25/27/28/29/30/31等）は、メニュー一覧と同じ商品ライン
        # であることが保証されている day_csv_paths を常に最優先で使う。
        # 「N月使用食材」シートは、DX弁当など別ラインのデータが同名パターンで紛れ込んでいる
        # ことがある（実データで確認済み）ため、day_csvが無い場合のみフォールバックで使う。
        if (month, '昼') in data.day_csv:
            data.shoku[month] = cm.build_day_index(data.day_csv[(month, '昼')])
        elif month in shoku_sheets:
            df = pd.read_excel(xl, shoku_sheets[month], header=0)
            df = _normalize_text_columns(df)
            data.shoku[month] = cm.build_day_index(df)
            data.warnings.append(f'{month}月：昼のday_csvが無いため「{month}月使用食材」シートを使用（別ラインの混在に注意）')
        else:
            data.warnings.append(f'{month}月：「{month}月使用食材」シートが見つからず、食材ベースの判定をスキップしました')
        if (month, '夜') in data.day_csv:
            data.shoku_night[month] = cm.build_day_index(data.day_csv[(month, '夜')])
        elif month in night_sheets:
            df = pd.read_excel(xl, night_sheets[month], header=0)
            df = _normalize_text_columns(df)
            data.shoku_night[month] = cm.build_day_index(df)
        elif month in night_csv_paths:
            try:
                df = pd.read_csv(night_csv_paths[month])
                df = _normalize_text_columns(df)
                data.shoku_night[month] = cm.build_day_index(df)
            except Exception as e:
                data.warnings.append(f'{month}月：夜食材CSVの読み込みに失敗しました（{e}）')
        else:
            data.warnings.append(f'{month}月：夜（夕）の食材データが見つからず、昼夜合算が必要な一部ルールをスキップしました')
        # No.14（栄養素基準）専用。優先順位：
        #   1) 「N月栄養価」シート（1日1行に集計済みのkcal/protein/salt。ユーザー指定）
        #   2) 「N月使用食材」シート（カロリー等の列を持つ場合、食材単位で合算）
        if month in nutrition_daily_sheets:
            df_daily = _find_nutrition_daily_sheet(wb_raw, nutrition_daily_sheets[month])
            if len(df_daily):
                data.nutrition_daily[month] = df_daily
            else:
                data.warnings.append(f'{month}月：「{nutrition_daily_sheets[month]}」からエネルギー列を検出できませんでした')
        elif month in shoku_sheets:
            df_n = pd.read_excel(xl, shoku_sheets[month], header=0)
            df_n = _normalize_text_columns(df_n)
            if all(c in df_n.columns for c in NUTRI_COLS):
                data.nutrition_shoku[month] = cm.build_day_index(df_n)
            else:
                data.warnings.append(f'{month}月：「{month}月使用食材」シートに栄養価列が無いため、No.14（栄養素基準）は{month}月をスキップしました')
        if month in lunchdinner_sheets:
            ws = wb_raw[lunchdinner_sheets[month]]
            data.rows += _parse_lunch_dinner_sheet(ws)
        else:
            data.warnings.append(f'{month}月：「{month}月昼夕...」シートが見つからず、No.24/28/29の判定をスキップしました')
        if month in lunch_tagged_sheets:
            df = pd.read_excel(xl, lunch_tagged_sheets[month], header=None)
            data.menu_lunch_tagged[month] = df
    all_dates = [r[0] for r in data.rows]
    if all_dates:
        data.date_range = pd.date_range(min(all_dates), max(all_dates))
    else:
        # rowsが無い場合はshokuの日付から範囲を作る
        mds = []
        for month, shoku in data.shoku.items():
            for md in shoku['md'].dropna().unique():
                try:
                    mds.append(datetime.date(2000 + (0 if month >= 1 else 0), int(md[0]), int(md[1])))
                except Exception:
                    pass
        if mds:
            data.date_range = pd.date_range(min(mds), max(mds))
    return data


def _shoku_for(data, month, night=False):
    d = data.shoku_night if night else data.shoku
    return d.get(month)


def raw_dish_names(data, date):
    names = set()
    month = date.month
    md = (date.month, date.day)
    for shoku in [data.shoku.get(month), data.shoku_night.get(month)]:
        if shoku is None:
            continue
        sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
        for _, r in sub.iterrows():
            prod = str(r['商品名'])
            qty = r.get('食材数量')
            if pd.isna(qty) or qty == 0 or cm.is_noise(prod):
                continue
            rn = str(r['レシピ名'])
            if rn and '備品' not in rn:
                names.add(rn)
    return names


def dishes_products(data, date):
    """(レシピ名 -> [商品名,...]) を昼夜合算で返す（基礎調味料含む）"""
    month = date.month
    md = (date.month, date.day)
    out = {}
    for shoku in [data.shoku.get(month), data.shoku_night.get(month)]:
        if shoku is None:
            continue
        sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
        for _, r in sub.iterrows():
            prod = str(r['商品名'])
            qty = r.get('食材数量')
            if pd.isna(qty) or qty == 0 or cm.is_noise(prod):
                continue
            out.setdefault(str(r['レシピ名']), []).append(prod)
    return out


def _min_gap_check(data, match_fn, min_gap, rule_no, rule_name, severity='中'):
    """スペーシング型：min_gap日以内の再使用はNG（例：かにのふわふわ5日以上空ける）"""
    dr = data.date_range
    dates_with = []
    for d in dr:
        names = raw_dish_names(data, d)
        hit = [n for n in names if match_fn(n)]
        if hit:
            dates_with.append((d, hit))
    viol = []
    for i in range(1, len(dates_with)):
        d0, _ = dates_with[i - 1]
        d1, h1 = dates_with[i]
        gap = (d1 - d0).days
        if gap <= min_gap:
            pool = _filtered_dish_hist(data, match_fn)
            cand = _pick_least_recent(pool.keys(), pool, d1, exclude={h1[0]})
            suggestion = f'代わりに「{cand[:20]}」等に変更' if cand else '使用日をずらす'
            viol.append({
                '日付': d1.strftime('%-m/%-d'), '曜日': WD_JP[d1.weekday()], 'No': rule_no, 'ルール': rule_name,
                '該当箇所': f'前回{d0.strftime("%-m/%-d")} → 今回{d1.strftime("%-m/%-d")}:{h1[0][:16]}',
                '理由': f'{gap}日しか空いていない（要{min_gap + 1}日以上）',
                '修正提案': suggestion, '重要度': severity,
            })
    return pd.DataFrame(viol)


def _max_gap_check(data, match_fn, max_gap, rule_no, rule_name, severity='中'):
    """頻度型：max_gap日を超えて使用が無いとNG（例：魚メニューは3日に1回以上）"""
    dr = data.date_range
    dates_with = []
    for d in dr:
        names = raw_dish_names(data, d)
        hit = [n for n in names if match_fn(n)]
        if hit:
            dates_with.append((d, hit))
    viol = []
    pool = _filtered_dish_hist(data, match_fn)
    if dates_with:
        gap0 = (dates_with[0][0] - dr[0]).days
        if gap0 > max_gap:
            cand = _pick_least_recent(pool.keys(), pool, dr[0])
            suggestion = f'「{cand[:20]}」等を追加検討' if cand else '追加を検討'
            viol.append({
                '日付': dr[0].strftime('%-m/%-d'), '曜日': WD_JP[dr[0].weekday()], 'No': rule_no,
                'ルール': rule_name + '（期間開始〜初回）',
                '該当箇所': f'{dr[0].strftime("%-m/%-d")}〜{dates_with[0][0].strftime("%-m/%-d")}',
                '理由': f'{gap0}日間使用なし', '修正提案': suggestion, '重要度': severity,
            })
    for i in range(1, len(dates_with)):
        d0, _ = dates_with[i - 1]
        d1, h1 = dates_with[i]
        gap = (d1 - d0).days
        if gap > max_gap:
            cand = _pick_least_recent(pool.keys(), pool, d1)
            suggestion = f'間隔内に「{cand[:20]}」等を追加' if cand else '間隔内に追加'
            viol.append({
                '日付': d1.strftime('%-m/%-d'), '曜日': WD_JP[d1.weekday()], 'No': rule_no, 'ルール': rule_name,
                '該当箇所': f'前回{d0.strftime("%-m/%-d")} → 今回{d1.strftime("%-m/%-d")}:{h1[0][:16]}',
                '理由': f'{gap}日間使用なし（上限{max_gap}日）',
                '修正提案': suggestion, '重要度': severity,
            })
    return pd.DataFrame(viol)


# ---------------- 各ルールの判定関数（run_check.py の最終版ロジックを月非依存化） ----------------

WEIGHT_UNITS = {'g', 'cc', 'ｇ', 'ｍｌ', 'ml'}
POS_ORDER_5 = ['メイン', 'サブ', '副菜1', '副菜2', 'サラダ']


def _usage_history(data):
    """商品名(NFKC)ごとに使用日(Timestamp)のソート済みリストを返す（day_csv全体から構築、キャッシュ有）。
    代替え案（NG時の具体的な代替商材提案）で『直近使われていない候補』を選ぶために使う。
    day_csvが渡されていない場合は、「N月使用食材」シート由来のdata.shoku/shoku_nightから
    同じ形の履歴を作る（day_csv無しでも代替え案を具体名で出せるようにするため）。"""
    if data._usage_hist is not None:
        return data._usage_hist
    md2date = {(ts.month, ts.day): ts for ts in (data.date_range if data.date_range is not None else [])}
    hist = {}
    sources = list(data.day_csv.values())
    if not sources:
        # フォールバック：day_csvが無い場合は使用食材シート（build_day_index済み）を使う
        sources = [df for df in list(data.shoku.values()) + list(data.shoku_night.values())
                   if df is not None]
    for df in sources:
        if df is None or '商品名' not in df.columns:
            continue
        for _, r in df.iterrows():
            name = str(r['商品名'])
            if cm.is_noise(name) or '終売' in name:
                continue
            md = r.get('md')
            if not isinstance(md, tuple):
                continue
            dt = md2date.get(md)
            if dt is None:
                continue
            hist.setdefault(name, set()).add(dt)
    data._usage_hist = {k: sorted(v) for k, v in hist.items()}
    return data._usage_hist


def _days_since_last_use(hist, name, before_date):
    """before_date時点で、その商品(name)が最後に使われたのが何日前か。未使用なら大きな値を返す。
    hist側はTimestamp、呼び出し側はdatetime.dateの場合があるため型を揃えて比較する。"""
    bd = pd.Timestamp(before_date)
    uses = [pd.Timestamp(d) for d in hist.get(name, [])]
    uses = [d for d in uses if d < bd]
    if not uses:
        return 10 ** 6
    return (bd - uses[-1]).days


def _pick_least_recent(candidates, hist, before_date, exclude=()):
    """candidates（商品名のiterable）の中から、before_date時点で最も長く使われていない
    （＝直近未使用の）ものを選んで返す。excludeに含まれるものは除外。"""
    best, best_gap = None, -1
    for c in candidates:
        if c in exclude:
            continue
        gap = _days_since_last_use(hist, c, before_date)
        if gap > best_gap:
            best, best_gap = c, gap
    return best


def _group_products_map(data):
    """主原料グループ(cm.group_from_name、パッチ適用済み) -> {商品名, ...} のマップ（キャッシュ有）。
    No.1の代替商材候補選定に使う。"""
    if getattr(data, '_group_products', None) is not None:
        return data._group_products
    hist = _usage_history(data)
    out = {}
    for name in hist:
        g = cm.group_from_name(name)
        out.setdefault(g, set()).add(name)
    data._group_products = out
    return out


def _dish_usage_history(data):
    """レシピ名(dish名) -> 使用日リスト（キャッシュ有）。No.4/36の代替案候補選定に使う。"""
    if getattr(data, '_dish_hist', None) is not None:
        return data._dish_hist
    hist = {}
    for d in data.date_range:
        for n in raw_dish_names(data, d):
            hist.setdefault(n, set()).add(d)
    data._dish_hist = {k: sorted(v) for k, v in hist.items()}
    return data._dish_hist


def _recipe_products(data):
    """レシピ名 -> そのレシピで使う {商品名, ...} のマップ（キャッシュ有）。
    代替え案を「商材」ではなく「レシピ（メニュー）」で出すために、
    『この食材/調味料を使っていないレシピ』を絞り込む用途で使う（ユーザー指定）。"""
    if getattr(data, '_recipe_prods', None) is not None:
        return data._recipe_prods
    out = {}
    sources = list(data.day_csv.values())
    if not sources:
        sources = [df for df in list(data.shoku.values()) + list(data.shoku_night.values())
                   if df is not None]
    for df in sources:
        if df is None or 'レシピ名' not in df.columns or '商品名' not in df.columns:
            continue
        for recipe, grp in df.groupby('レシピ名', sort=False):
            rn = str(recipe)
            if '備品' in rn:
                continue
            out.setdefault(rn, set()).update(grp['商品名'].astype(str).tolist())
    data._recipe_prods = out
    return out


def _recipe_product_ids(data):
    """レシピ名 -> そのレシピで使う {商品ID(int), ...} のマップ（キャッシュ有）。No.30等のID判定用。"""
    if getattr(data, '_recipe_pids', None) is not None:
        return data._recipe_pids
    out = {}
    sources = list(data.day_csv.values())
    if not sources:
        sources = [df for df in list(data.shoku.values()) + list(data.shoku_night.values())
                   if df is not None]
    for df in sources:
        if df is None or 'レシピ名' not in df.columns or '商品ID' not in df.columns:
            continue
        for recipe, grp in df.groupby('レシピ名', sort=False):
            rn = str(recipe)
            if '備品' in rn:
                continue
            ids = pd.to_numeric(grp['商品ID'], errors='coerce').dropna().astype(int)
            out.setdefault(rn, set()).update(ids.tolist())
    data._recipe_pids = out
    return out


def _recipe_has(data, recipe, kw):
    """そのレシピが、商品名またはレシピ名に kw を含むか（NFKC正規化して判定）"""
    if kw in _nfkc(recipe):
        return True
    return any(kw in _nfkc(p) for p in _recipe_products(data).get(recipe, ()))


def _recipe_replacement(data, date, ok=None, group=None, exclude=()):
    """レシピ（メニュー）単位の代替え案を1つ返す。候補が無ければ None。
    ・ok    : そのレシピ名を候補にしてよいか判定する関数（Noneなら全て可）
    ・group : 同系統（cm.group_from_name）を優先したい場合に指定
    ・exclude: 除外するレシピ名
    いずれも『その日時点で最も長く使われていないレシピ』を選ぶ。
    ユーザー指定により、代替え案は原則すべて商材名ではなくレシピ名で出す。"""
    return _recipe_replacement2(data, date, ok=ok, group=group, exclude=exclude)[0]


def _recipe_replacement2(data, date, ok=None, group=None, exclude=()):
    """_recipe_replacement の (レシピ名, 同系統で見つかったか) を返す版。
    「同系統（◯◯）の…」という文言を出してよいかを呼び出し側が判断できるようにするため。
    主原料グループ '他' は寄せ集めのため、同系統扱いにはしない。"""
    hist = _dish_usage_history(data)
    if not hist:
        return None, False
    ex = set(exclude)
    cands = [n for n in hist if n not in ex and (ok is None or ok(n))]
    if not cands:
        return None, False
    if group and group != '他':
        same = [n for n in cands if cm.group_from_name(n) == group]
        pick = _pick_least_recent(same, hist, date)
        if pick:
            return pick, True
    return _pick_least_recent(cands, hist, date), False


def _nonfried_dish_names(data):
    """当日揚げ（data.fried_recipe_ids）に該当しないレシピ名の集合（キャッシュ有）。
    No.12（揚げ物超過）の『非揚げ物への差し替え』代替案候補に使う。"""
    if getattr(data, '_nonfried_names', None) is not None:
        return data._nonfried_names
    fried_names = set()
    all_names = set()
    for df in data.day_csv.values():
        sub = df[~df['レシピ名'].astype(str).str.contains('備品', na=False)]
        for recipe, grp in sub.groupby('レシピ名', sort=False):
            all_names.add(str(recipe))
            ids = grp['レシピID'].dropna().astype(int)
            if ids.isin(data.fried_recipe_ids).any():
                fried_names.add(str(recipe))
    data._nonfried_names = all_names - fried_names
    return data._nonfried_names


def _nutrition_candidates(data, column, ascending=False, top=3):
    """レシピ別栄養価マスタ(recipe_nutrition)から、指定栄養素の多い順（または少ない順）に
    レシピ名を返す。No.14の「具体的にどのメニューを足す/差し替えるか」の提案に使う。"""
    df = data.recipe_nutrition
    if df is None or not len(df) or column not in df.columns:
        return []
    sub = df.dropna(subset=[column]).sort_values(column, ascending=ascending)
    names = []
    for nm in sub['名称'].astype(str):
        nm = nm.strip()
        if nm and nm not in names:
            names.append(nm)
        if len(names) >= top:
            break
    return names


def _next_weekday(d, max_ahead=7):
    """dの翌日以降で最初の平日（月〜金）を返す。見つからなければNone。
    No.27/28の『いつの平日枠に振り替えるか』の具体案に使う。"""
    d0 = pd.Timestamp(d)
    for i in range(1, max_ahead + 1):
        nd = d0 + pd.Timedelta(days=i)
        if nd.weekday() < 5:
            return nd
    return None


def _flexible_veg_names(data):
    """VEG_FLEXIBLE_IDS（間隔制約がほぼ無い野菜）の実データ上の商品名一覧（キャッシュ有）。
    No.30違反時の『代わりに使える野菜』提案に使う。"""
    if getattr(data, '_flex_veg', None) is not None:
        return data._flex_veg
    names = set()
    ids = set(VEG_FLEXIBLE_IDS)
    for df in data.day_csv.values():
        hit = df[pd.to_numeric(df['商品ID'], errors='coerce').isin(ids)]
        names |= set(hit['商品名'].astype(str))
    data._flex_veg = sorted(n for n in names if '終売' not in n)
    return data._flex_veg


def _ng_replacement(data, prod_name, date, recipe_name=None, ng_words=()):
    """No.21用：禁止食材・調味料に該当したメニューの代替え案を『レシピ（メニュー）名』で返す。
    ユーザー指定により、調味料（タレ/ソース/〜の素）が禁止対象の場合も、調味料の差し替えでは
    なくレシピごと差し替える案を出す。
    候補は「禁止対象（商品名・禁止ワード）を含まないレシピ」で、同系統(cm.group_from_name)を
    優先し、その日時点で最も長く使われていないものを選ぶ。"""
    words = [w for w in (list(ng_words) or NG_WORDS)]

    def _safe(n):
        if any(w in str(n) for w in words):
            return False
        if prod_name and prod_name in _recipe_products(data).get(n, ()):
            return False
        if any(_recipe_has(data, n, w) for w in words):
            return False
        return True

    base = recipe_name or prod_name or ''
    group = cm.group_from_name(base)
    cand, same_group = _recipe_replacement2(data, date, ok=_safe, group=group, exclude={base})
    if cand and same_group:
        return f'同系統（{group}）の「{cand[:26]}」に差し替え'
    if cand:
        return f'「{cand[:26]}」等、禁止食材を含まないメニューに差し替え'
    return '禁止食材を含まないメニューに差し替え'


def _filtered_dish_hist(data, match_fn):
    """_dish_usage_history() のうち、match_fn(name)がTrueのものだけに絞ったdict（キャッシュ無し・軽量）。
    「頻度/間隔系」ルール（No.15/22/25/26/29等）で、追加すべき具体的な代替候補を選ぶのに使う。"""
    return {k: v for k, v in _dish_usage_history(data).items() if match_fn(k)}


def veg_names_for(name, veg_color_map):
    """商品名/レシピ名にマッチする野菜マスタの正規化名（複数ヒットしうる）を返す"""
    n = _nfkc(str(name))
    return [veg_name for veg_name, _c in veg_color_map if veg_name in n]


def _veg_usage_history(data):
    """野菜マスタの正規化名 -> 使用日リスト（キャッシュ有）。No.9/No.17の代替案候補選定に使う。"""
    if getattr(data, '_veg_hist', None) is not None:
        return data._veg_hist
    hist = {}
    prod_hist = _usage_history(data)
    for name, dates in prod_hist.items():
        for vn in veg_names_for(name, data.veg_color_map):
            hist.setdefault(vn, set()).update(dates)
    data._veg_hist = {k: sorted(v) for k, v in hist.items()}
    return data._veg_hist


def _day_recipe_order(day_csv, month, day, slot):
    """食材CSV上で、その日・その時間帯（昼/夜）に登場するレシピ名を初出順に並べる。
    （メニュー一覧シートの並び=メイン→サブ→副菜1→副菜2→サラダ と、食材CSV内の
    レシピ出現順が一致することを実データで確認済み。備品行は除外）"""
    df = day_csv.get((month, slot))
    if df is None:
        return []
    sub = df[df['md'] == (month, day)]
    seen = []
    for rn in sub['レシピ名'].astype(str):
        if '備品' in rn:
            continue
        if rn not in seen:
            seen.append(rn)
    return seen


def _primary_product(day_csv, month, day, slot, recipe):
    """そのレシピの中で『個数カウント（個/枚/尾等）の単体商材』を1つ返す。
    g/cc計量の生食材・調味料を組み合わせただけの料理（単体商材が無い/複数ある）はNoneを返し、
    No.1の判定対象外とする（ユーザー確認済みの運用ルール）。"""
    df = day_csv.get((month, slot))
    sub = df[(df['md'] == (month, day)) & (df['レシピ名'] == recipe)]
    sub = sub[~sub['商品名'].astype(str).apply(cm.is_noise)]
    piece = sub[~sub['ユニット名'].isin(WEIGHT_UNITS)]
    ids = piece['商品ID'].dropna().unique()
    if len(ids) == 1:
        pname = piece[piece['商品ID'] == ids[0]]['商品名'].iloc[0]
        return ids[0], str(pname)
    return None, None


def check_rule1(data):
    """No.1: 同一商品ID（個数カウントの単体商材＝コロッケ/カツ/フライ等の冷凍加工品）を
    メイン・サブメインで1週間以内（8日未満）に再使用したらNG。
    料理名（レシピ名/メニュー名）が違っていても、中身の商品が同じなら検出する。
    g/cc計量の生食材を組み合わせた料理（単体商材が無いもの）は対象外。
    data.day_csv: {(month, '昼'/'夜'): DataFrame} 形式の食材CSV（'md'列で日付紐付け済み）が
    必要。未設定の場合は判定不能のため空のDataFrameを返す。"""
    day_csv = data.day_csv
    if not data.rows or not day_csv:
        return pd.DataFrame()
    seen_days = sorted(set((d, slot) for (d, wd, slot, pos, name) in data.rows))
    entries = []
    for (d, slot) in seen_days:
        order = _day_recipe_order(day_csv, d.month, d.day, slot)
        for i, pos_label in enumerate(('メイン', 'サブ')):
            if i >= len(order):
                break
            recipe = order[i]
            pid, pname = _primary_product(day_csv, d.month, d.day, slot, recipe)
            if pid is not None:
                entries.append((d, slot, pos_label, recipe, pid, pname))
    entries.sort(key=lambda x: x[0])
    last_seen = {}
    viol = []
    group_map = _group_products_map(data)
    hist = _usage_history(data)
    for d, slot, pos, recipe, pid, pname in entries:
        if pid in last_seen:
            prev_d, prev_slot, prev_pos, prev_recipe, prev_pname = last_seen[pid]
            gap = (d - prev_d).days
            if 0 < gap <= 7:
                group = cm.group_from_name(pname)
                # 代替え案はレシピ（メニュー）名で出す：同系統で、その商材を使っていないレシピ
                cand, same_group = _recipe_replacement2(
                    data, d, ok=lambda n: pname not in _recipe_products(data).get(n, ()),
                    group=group, exclude={recipe, prev_recipe})
                if cand and same_group:
                    suggestion = f'同系統（{group}）の「{cand[:26]}」に変更を検討'
                elif cand:
                    suggestion = f'「{cand[:26]}」等、{pname[:14]}を使わないメニューに変更'
                else:
                    suggestion = '該当日か次回使用日をずらす'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 1,
                    'ルール': '同一商品（単体商材）をメイン/サブで1週間空けず再使用',
                    '該当箇所': f'{slot}{pos}:{recipe[:20]}（商品:{pname[:22]}）← 前回 {prev_d.strftime("%-m/%-d")} {prev_slot}{prev_pos}:{prev_recipe[:18]}',
                    '理由': f'同一商品ID（{int(pid)}）を{gap}日しか空けず再使用（要8日以上）',
                    '修正提案': suggestion, '重要度': '高',
                })
        last_seen[pid] = (d, slot, pos, recipe, pname)
    return pd.DataFrame(viol)


def _visual_group(name):
    """LOOKALIKE_PAIRSに登録された「見た目が同じ」グループの代表名を返す。未登録ならNone。
    （No.2用。No.1のcanon_nameと違い、フレームタイトル処理はせず単純に部分一致で見る）"""
    n = str(name)
    for group, _gap in LOOKALIKE_PAIRS:
        for g in group:
            if g in n:
                return group[0]
    return None


def check_rule2(data, max_per_month=2, min_gap_days=8):
    """No.2: メイン/サブで『見た目が同じもの』は1週間以上空けて使用する（月2回まで）。
    例：海老カツ、海老タラカツ（LOOKALIKE_PAIRSに登録された組を「見た目グループ」とみなす）。
    No.1（同一商品ID）とは別軸で、実際の商品が違っても見た目の系統が同じなら対象になる。
    現状 LOOKALIKE_PAIRS は4組のみ登録（実データを確認して随時追加する運用）のため、
    ここに登録されていない『見た目が同じ』ケースは検出できない点に注意。"""
    entries = sorted(
        [(d, slot, pos, name) for (d, wd, slot, pos, name) in data.rows if pos in ('メイン', 'サブ')],
        key=lambda x: x[0],
    )
    viol = []
    last_seen = {}   # group -> (date, slot, pos, name)
    month_count = {}  # (group, year, month) -> count
    for d, slot, pos, name in entries:
        group = _visual_group(name)
        if not group:
            continue
        # 1) 1週間（8日）以上空いているか
        if group in last_seen:
            prev_d, prev_slot, prev_pos, prev_name = last_seen[group]
            gap = (d - prev_d).days
            if 0 < gap < min_gap_days:
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 2,
                    'ルール': 'メイン/サブで見た目が同じ商材を1週間空けず再使用',
                    '該当箇所': f'{slot}{pos}:{name[:20]}（見た目グループ:{group}）← 前回 {prev_d.strftime("%-m/%-d")} {prev_slot}{prev_pos}:{prev_name[:16]}',
                    '理由': f'{gap}日しか空けず再使用（要{min_gap_days}日以上）',
                    '修正提案': '該当日か次回使用日をずらす', '重要度': '高',
                })
        last_seen[group] = (d, slot, pos, name)
        # 2) 月2回まで
        key = (group, d.year, d.month)
        month_count[key] = month_count.get(key, 0) + 1
        if month_count[key] > max_per_month:
            viol.append({
                '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 2,
                'ルール': 'メイン/サブで見た目が同じ商材が月2回を超過',
                '該当箇所': f'{slot}{pos}:{name[:20]}（見た目グループ:{group}）',
                '理由': f'{d.month}月内で{month_count[key]}回目（上限{max_per_month}回）',
                '修正提案': '翌月以降にずらす', '重要度': '中',
            })
    return pd.DataFrame(viol)


def check_rule3_5(data):
    """No.3(挽肉重複) / No.5(鶏豚牛重複、No.34のハンバーグ例外込み)（昼夜それぞれの食事単位で判定）"""
    if not data.rows:
        return pd.DataFrame(), pd.DataFrame()
    by_ds = {}
    for (d, wd, slot, pos, name) in data.rows:
        if pos in ('メイン', 'サブ'):
            by_ds.setdefault((d, slot), {})[pos] = name
    v3, v5 = [], []
    for (d, slot), posmap in sorted(by_ds.items()):
        nm_m, nm_s = posmap.get('メイン', ''), posmap.get('サブ', '')
        if not nm_m or not nm_s:
            continue
        gm, gs = cm.group_from_name(nm_m), cm.group_from_name(nm_s)
        if gm != gs:
            continue
        dish_hist = _dish_usage_history(data)
        if gm == 'ひき肉系':
            is_exception = ('豆腐ハンバーグ' in nm_m and 'ハンバーグ' in nm_s and '豆腐ハンバーグ' not in nm_s) or \
                            ('豆腐ハンバーグ' in nm_s and 'ハンバーグ' in nm_m and '豆腐ハンバーグ' not in nm_m)
            if is_exception:
                continue
            cand = _pick_least_recent(
                [n for n in dish_hist if cm.group_from_name(n) != gm], dish_hist, d, exclude={nm_m, nm_s})
            suggestion = f'サブを「{cand[:18]}」等、別系統に変更' if cand else 'メインかサブの系統を変える'
            v3.append({
                '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 3,
                'ルール': '挽肉商材がメイン・サブメインで同日重複',
                '該当箇所': f'[{slot}] メイン:{nm_m[:16]} / サブ:{nm_s[:16]}', '理由': f'両方「{gm}」',
                '修正提案': suggestion, '重要度': '高',
            })
        if gm in ('鶏肉系', '豚肉系', '牛肉系'):
            cand = _pick_least_recent(
                [n for n in dish_hist if cm.group_from_name(n) not in ('鶏肉系', '豚肉系', '牛肉系')],
                dish_hist, d, exclude={nm_m, nm_s})
            suggestion = f'サブを「{cand[:18]}」等、別系統に変更' if cand else 'メインかサブの系統を変える'
            v5.append({
                '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 5,
                'ルール': '鶏豚牛が同日でメイン・サブメインに重複（枠をずらす）',
                '該当箇所': f'[{slot}] メイン:{nm_m[:16]} / サブ:{nm_s[:16]}', '理由': f'両方「{gm}」',
                '修正提案': suggestion, '重要度': '高',
            })
    return pd.DataFrame(v3), pd.DataFrame(v5)


def check_rule4_36(data):
    """No.4(コロッケ1日空け) / No.36(通常⇔クリームは連日可、の例外込み)"""
    dr = data.date_range
    hits = []
    for d in dr:
        names = raw_dish_names(data, d)
        for n in names:
            if 'コロッケ' in n:
                cat = 'クリーム' if 'クリーム' in n else '通常'
                hits.append((d, n, cat))
    hits.sort(key=lambda x: x[0])
    viol = []
    last_seen = {}
    dish_hist = _dish_usage_history(data)
    for d, n, cat in hits:
        if cat in last_seen:
            pd0, pn0 = last_seen[cat]
            gap = (d - pd0).days
            # 同日内の複数サイズ登録（例：S用30g・M用50g）は重複とみなさない（ユーザー確認済み）。
            # gap==1（翌日）のみ「連日重複」としてNG。
            if gap == 1:
                candidates = [n2 for n2 in dish_hist if 'コロッケ' in n2 and
                              (('クリーム' in n2) == (cat == 'クリーム'))]
                cand = _pick_least_recent(candidates, dish_hist, d, exclude={n})
                suggestion = f'別のコロッケ「{cand[:20]}」に変更を検討' if cand else '使用日をずらす'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 4,
                    'ルール': f'コロッケ({cat})が連日重複',
                    '該当箇所': f'{n[:16]}（前回 {pd0.strftime("%-m/%-d")} {pn0[:16]}）',
                    '理由': f'{gap}日しか空いていない', '修正提案': suggestion, '重要度': '中',
                })
        last_seen[cat] = (d, n)
    return pd.DataFrame(viol)


# No.8で「調味料の中身」として見る食材ワード。
# 調味料.csv の商品名に含まれる“素材名”のうち、同じ食材が生食材としても使われうるものだけを挙げる。
# 塩・こしょう・砂糖・醤油・酢・油のような、ほぼ全料理に入る基礎調味料の要素は
# 検出してもきりが無いため入れない（ユーザー確認済み）。
# 玉ねぎ(オニオン)・にんにく等の基礎野菜は、塩こしょうと同じくほぼ全料理で使われ
# 検出してもきりが無いため入れない（ユーザー確認済み。BASE_VEG_KWと同じ考え方）。
SEASONING_FOOD_KW = [
    '生姜', 'しょうが', 'ごま', '胡麻', '梅', '柚子', 'ゆず', '青じそ', '大葉', '昆布',
    'かつお', 'トマト', 'バジル', 'アンチョビ', '豆乳', 'ピーナッツ', '大根おろし',
    'レモン', 'マスタード', 'チーズ', 'バター', 'わさび', 'ゆかり', 'しそ',
]


def check_rule8(data):
    """No.8: 1食のうち『食材と調味料での食材被り』はNG。
    例：生姜おろし（食材）と やどかり国産生姜焼きのタレ（調味料）を同じ食事内で使う。
    調味料かどうかは調味料.csv（商品ID）で判定し、その商品名に含まれる素材名
    （SEASONING_FOOD_KW）が、同じ食事の“調味料でない食材”にも含まれていればNGとする。
    塩・こしょう・醤油・砂糖などの基礎調味料要素は、ほぼ全料理で使われ検出してもきりが無いため
    SEASONING_FOOD_KWに含めない（ユーザー確認済み）。
    昼/夜は別々の「1食」として判定する。"""
    if not data.day_csv or not data.seasoning_ids:
        return pd.DataFrame()
    dr = data.date_range
    viol = []
    for d in dr:
        for slot in ('昼', '夜'):
            df = data.day_csv.get((d.month, slot))
            if df is None:
                continue
            sub = df[(df['md'] == (d.month, d.day)) &
                     (~df['レシピ名'].astype(str).str.contains('備品', na=False))]
            if not len(sub):
                continue
            ids = pd.to_numeric(sub['商品ID'], errors='coerce')
            seasonings = sub[ids.isin(data.seasoning_ids)].drop_duplicates('商品ID')
            foods = sub[~ids.isin(data.seasoning_ids)]
            seen = set()
            for _, sr in seasonings.iterrows():
                sname = str(sr['商品名'])
                sname_n = _nfkc(sname)
                if cm.is_noise(sname):
                    continue
                for kw in SEASONING_FOOD_KW:
                    if kw not in sname_n:
                        continue
                    hit = foods[foods['商品名'].astype(str).apply(lambda x: kw in _nfkc(x))]
                    # 同一レシピ内での被り（例：豚肉生姜焼きに生姜おろし＋生姜焼きのタレ）は
                    # 料理として自然なため対象外。別レシピ間の被りのみをNGとする（ユーザー確認済み）。
                    hit = hit[hit['レシピ名'].astype(str) != str(sr['レシピ名'])]
                    if not len(hit):
                        continue
                    fname = str(hit['商品名'].iloc[0])
                    key = (slot, kw)
                    if key in seen:
                        continue
                    seen.add(key)
                    frecipe = str(hit['レシピ名'].iloc[0])
                    srecipe = str(sr['レシピ名'])
                    # 代替え案はレシピ名で出す：その素材（kw）を含まない別メニュー
                    cand = _recipe_replacement(
                        data, d, ok=lambda n: not _recipe_has(data, n, kw),
                        exclude={frecipe, srecipe})
                    suggestion = f'「{srecipe[:16]}」を「{cand[:24]}」等、{kw}を含まないメニューに差し替え' if cand \
                        else f'どちらかを{kw}を含まないメニューに変更'
                    viol.append({
                        '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 8,
                        'ルール': '1食内で食材と調味料の中身が被っている',
                        '該当箇所': f'[{slot}] {kw}：{fname[:20]}（{frecipe[:14]}） × {sname[:22]}（{srecipe[:14]}）',
                        '理由': f'食材と調味料の両方に「{kw}」が含まれる',
                        '修正提案': suggestion, '重要度': '低',
                    })
                    break
    return pd.DataFrame(viol)


COMMON_VEG_COLORS = {'緑', '黄', '白'}  # ほぼ毎食使われる基礎色（No.6/8の基礎調味料・基礎野菜と同じ考え方で除外）
# 赤・茶に分類されるが、彩り確保のためほぼ毎日使われる想定の定番食材（ユーザー確認済み・除外）
COMMON_VEG_NAMES = ['人参', 'にんじん', 'しいたけ', '椎茸']


def check_rule9(data):
    """No.9: 昼夕それぞれ、見た目（色）が同じ野菜が2日連続で使われたらNG。
    野菜マスタ_テンプレート.xlsx の「色（彩り）」列（ユーザー指定）で判定する。
    昼は昼同士、夜は夜同士で前日と比較する（日をまたいだ昼→夜比較はしない）。
    マスタに登録の無い野菜（色が未入力/未登録）は判定対象外。
    緑・黄・白はほぼ毎日どこかに使われる基礎色のため対象外とし（No.6/8と同じ方針）、
    彩りとして目立つ赤・紫・茶などの重複のみを検出する。"""
    if not data.veg_color_map:
        return pd.DataFrame()
    dr = data.date_range
    viol = []
    for slot, shoku_dict in [('昼', data.shoku), ('夜', data.shoku_night)]:
        prev_date = None
        prev_colors = {}  # 色 -> 商品名(代表1つ)
        for d in dr:
            month = d.month
            shoku = shoku_dict.get(month)
            if shoku is None:
                continue
            md = (d.month, d.day)
            sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
            today_colors = {}
            for _, r in sub.iterrows():
                prod = str(r['商品名'])
                qty = r.get('食材数量')
                if pd.isna(qty) or qty == 0 or cm.is_noise(prod):
                    continue
                if any(k in _nfkc(prod) for k in COMMON_VEG_NAMES):
                    continue
                for c in veg_colors_for(prod, data.veg_color_map):
                    if c in COMMON_VEG_COLORS:
                        continue
                    today_colors.setdefault(c, prod)
            if prev_date is not None and (d - prev_date).days == 1:
                overlap = set(today_colors) & set(prev_colors)
                for c in overlap:
                    # 代替え案はレシピ名で出す：その色の野菜を含まないメニュー
                    def _no_color(n, _c=c):
                        prods = _recipe_products(data).get(n, ())
                        cols = set()
                        for p in prods:
                            cols |= veg_colors_for(p, data.veg_color_map)
                        return bool(cols) and _c not in cols
                    cand = _recipe_replacement(data, d, ok=_no_color)
                    suggestion = f'「{cand[:26]}」等、{c}以外の色の野菜メニューに変更を検討' if cand \
                        else f'{c}以外の色の野菜メニューに変更'
                    viol.append({
                        '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 9,
                        'ルール': '見た目（色）が同じ野菜が2日連続',
                        '該当箇所': f'[{slot}] {c}：{today_colors[c][:20]}（前日 {prev_date.strftime("%-m/%-d")}：{prev_colors[c][:20]}）',
                        '理由': f'同じ色（{c}）の野菜が前日と連続', '修正提案': suggestion, '重要度': '低',
                    })
            if today_colors:
                prev_date, prev_colors = d, today_colors
    return pd.DataFrame(viol)


def check_rule10(data, min_gap_days=8):
    """No.10: 同一食材のみで構成したメニュー（副菜・サラダ）は、味付けを変えても
    1週間以上空けて使用する。day_csv（商品ID紐付けCSV）が必要（No.1と同じ仕組みを流用）。
    副菜1/副菜2/サラダの各レシピについて、基礎調味料を除いた食材が1品だけのものを
    『単一食材メニュー』とみなし、その食材の使用間隔を見る（味付けが変わっていても対象）。"""
    day_csv = data.day_csv
    if not day_csv or not data.rows:
        return pd.DataFrame()
    entries = []
    seen_days = sorted(set((d, slot) for (d, wd, slot, pos, name) in data.rows))
    for d, slot in seen_days:
        df = day_csv.get((d.month, slot))
        if df is None:
            continue
        order = _day_recipe_order(day_csv, d.month, d.day, slot)
        for i, pos_label in enumerate(POS_ORDER_5):
            if pos_label not in ('副菜1', '副菜2', 'サラダ') or i >= len(order):
                continue
            recipe = order[i]
            sub = df[(df['md'] == (d.month, d.day)) & (df['レシピ名'] == recipe)]
            prods = []
            for _, r in sub.iterrows():
                prod = str(r['商品名'])
                qty = r.get('食材数量')
                if pd.isna(qty) or qty == 0 or cm.is_noise(prod) or is_base_seasoning(prod):
                    continue
                prods.append(prod)
            uniq = list(dict.fromkeys(prods))
            if len(uniq) == 1:
                key = cm.norm_recipe(uniq[0]) or uniq[0]
                entries.append((d, slot, pos_label, recipe, key, uniq[0]))
    entries.sort(key=lambda x: x[0])
    # key(単一食材) -> [(date, recipe), ...]  代替案候補（他の単一食材メニュー）選定に使う
    key_dates = {}
    key_recipe = {}
    for d, slot, pos, recipe, key, prod in entries:
        key_dates.setdefault(key, []).append(d)
        key_recipe.setdefault(key, recipe)
    last_seen = {}
    viol = []
    for d, slot, pos, recipe, key, prod in entries:
        if key in last_seen:
            prev_d, prev_slot, prev_pos, prev_recipe, prev_prod = last_seen[key]
            gap = (d - prev_d).days
            if 0 < gap < min_gap_days:
                best_key, best_gap = None, -1
                for k2, dates2 in key_dates.items():
                    if k2 == key:
                        continue
                    past = [dt for dt in dates2 if dt < d]
                    g2 = (d - past[-1]).days if past else 10 ** 6
                    if g2 > best_gap:
                        best_key, best_gap = k2, g2
                suggestion = f'別の単一食材メニュー「{key_recipe[best_key][:18]}」に変更を検討' if best_key else '該当日か次回使用日をずらす'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 10,
                    'ルール': '単一食材のみの副菜/サラダを1週間空けず再使用',
                    '該当箇所': f'{slot}{pos}:{recipe[:18]}（食材:{prod[:16]}）← 前回 {prev_d.strftime("%-m/%-d")} {prev_slot}{prev_pos}:{prev_recipe[:16]}',
                    '理由': f'{gap}日しか空けず再使用（要{min_gap_days}日以上）', '修正提案': suggestion, '重要度': '低',
                })
        last_seen[key] = (d, slot, pos, recipe, prod)
    return pd.DataFrame(viol)


def check_rule11(data):
    """No.11: 1食1メニューは自然解凍品。商品名に「自然解凍」を含む商材が
    その食事（昼/夜）に1品も無ければNG（ユーザー指定：商品名の「自然解凍」表記で判定）。"""
    dr = data.date_range
    viol = []
    for d in dr:
        month = d.month
        for meal, shoku in [('昼', data.shoku.get(month)), ('夜', data.shoku_night.get(month))]:
            if shoku is None:
                continue
            md = (d.month, d.day)
            sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
            if not len(sub):
                continue
            has_natural = sub['商品名'].astype(str).str.contains('自然解凍', na=False).any()
            if not has_natural:
                # 代替え案はレシピ名で出す：自然解凍品を使っているメニュー
                cand = _recipe_replacement(data, d, ok=lambda n: _recipe_has(data, n, '自然解凍'))
                suggestion = f'副菜等を「{cand[:26]}」等、自然解凍品を使うメニューに変更' if cand \
                    else '副菜等を自然解凍品のメニューに変更'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': meal, 'No': 11,
                    'ルール': '1食1メニューは自然解凍品',
                    '該当箇所': f'[{meal}]', '理由': '商品名に「自然解凍」を含む商材が0品',
                    '修正提案': suggestion, '重要度': '中',
                })
    return pd.DataFrame(viol)


def check_rule6(data):
    """No.6: 1食内（昼/夜は別々）で同一食材が複数レシピに重複使用（基礎調味料・基礎野菜は除外）"""
    dr = data.date_range
    viol = []
    for d in dr:
        month = d.month
        for meal, shoku in [('昼', data.shoku.get(month)), ('夜', data.shoku_night.get(month))]:
            if shoku is None:
                continue
            md = (d.month, d.day)
            sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
            prod_recipes = {}
            for _, r in sub.iterrows():
                prod = str(r['商品名'])
                qty = r.get('食材数量')
                if pd.isna(qty) or qty == 0 or cm.is_noise(prod) or is_base_seasoning(prod) or is_base_veg(prod):
                    continue
                recipe = cm.norm_recipe(r['レシピ名'])
                if not recipe or '備品' in recipe:
                    continue
                prod_recipes.setdefault(prod, set()).add(recipe)
            for prod, recipes in prod_recipes.items():
                if len(recipes) >= 2:
                    # 代替え案はレシピ名で出す：その食材を使っていない別レシピ
                    cand = _recipe_replacement(
                        data, d, ok=lambda n: prod not in _recipe_products(data).get(n, ()),
                        exclude=set(recipes))
                    suggestion = f'一方を「{cand[:26]}」等、{prod[:14]}を使わないメニューに変更' if cand \
                        else 'いずれかを別食材のメニューに変更'
                    viol.append({
                        '日付': d.strftime('%-m/%-d'), '曜日': meal, 'No': 6,
                        'ルール': '1食内で同一食材が複数レシピに重複使用',
                        '該当箇所': f'[{meal}] {prod[:26]} → ' + ' / '.join(list(recipes)[:4]),
                        '理由': f'{len(recipes)}レシピで使用', '修正提案': suggestion, '重要度': '中',
                    })
    return pd.DataFrame(viol)


def raw_dish_names_slot(data, date, slot):
    """raw_dish_namesの昼/夜を分けた版。指定slot（'昼' or '夜'）のレシピ名だけを返す。"""
    names = set()
    month = date.month
    shoku = data.shoku.get(month) if slot == '昼' else data.shoku_night.get(month)
    if shoku is None:
        return names
    md = (date.month, date.day)
    sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
    for _, r in sub.iterrows():
        prod = str(r['商品名'])
        qty = r.get('食材数量')
        if pd.isna(qty) or qty == 0 or cm.is_noise(prod):
            continue
        rn = str(r['レシピ名'])
        if rn and '備品' not in rn:
            names.add(rn)
    return names


def check_rule7(data):
    """No.7: 大豆系商材は半日空ける＝同じ食事（昼は昼同士、夜は夜同士）内での複数使用はNG。
    昼と夜に分かれていれば半日以上空いているとみなし対象外とする（ユーザー確認済み）。"""
    dr = data.date_range
    viol = []
    for d in dr:
        for slot in ('昼', '夜'):
            names = raw_dish_names_slot(data, d, slot)
            soy = sorted(n for n in names if is_soy(n))
            if len(soy) >= 2:
                non_soy_hist = _filtered_dish_hist(data, lambda n: not is_soy(n))
                cand = _pick_least_recent(non_soy_hist.keys(), non_soy_hist, d)
                suggestion = f'一方を非大豆系の「{cand[:20]}」等に変更' if cand else '一方を別の食事にずらす'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 7,
                    'ルール': '大豆系商材が同じ食事内で複数使用（半日未満）',
                    '該当箇所': f'[{slot}] ' + ' / '.join(soy), '理由': f'{slot}に大豆系が{len(soy)}品',
                    '修正提案': suggestion, '重要度': '中',
                })
    return pd.DataFrame(viol)


def check_rule12(data):
    """No.12: 当日揚げ調理は3品まで。
    食材データ.xlsx「調理法（当日揚げ）」シート（レシピID基準）で判定する（ユーザー指定）。
    day_csvのレシピIDと突き合わせるため、調理法タグの有無に関わらず7月・8月どちらも判定できる。"""
    if not data.fried_recipe_ids or not data.day_csv:
        return pd.DataFrame()
    viol = []
    dr = data.date_range
    for d in dr:
        for slot in ('昼', '夜'):
            df = data.day_csv.get((d.month, slot))
            if df is None:
                continue
            sub = df[(df['md'] == (d.month, d.day)) & (~df['レシピ名'].astype(str).str.contains('備品', na=False))]
            recipe_ids = sub.dropna(subset=['レシピID'])[['レシピID', 'レシピ名']].drop_duplicates()
            fried = recipe_ids[recipe_ids['レシピID'].astype(int).isin(data.fried_recipe_ids)]
            if len(fried) >= 4:
                names = '/'.join(fried['レシピ名'].astype(str).str[:12].tolist())
                nonfried_hist = _filtered_dish_hist(data, lambda n: n in _nonfried_dish_names(data))
                cand = _pick_least_recent(nonfried_hist.keys(), nonfried_hist, d)
                suggestion = f'いずれか1品を「{cand[:18]}」等の非揚げ物に変更' if cand else '1品を煮/和え等に'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 12,
                    'ルール': '当日揚げ調理が3品を超過', '該当箇所': f'[{slot}] {names}',
                    '理由': f'当日揚げが{len(fried)}品（上限3品）',
                    '修正提案': suggestion, '重要度': '中',
                })
    return pd.DataFrame(viol)


def check_rule14(data):
    """No.14: 75歳以上向け栄養素基準（月平均・75歳基準、昼夜それぞれの基準値で判定）。
    優先的に nutrition_daily（「N月栄養価」シート・1日1行で集計済み。ユーザー指定）を使う。
    無い月のみ nutrition_shoku（「N月使用食材」シート・食材単位で合算）にフォールバックする。
    夜は現状、栄養価データが無いためスキップされる（data.warningsに記録）。"""
    viol = []
    slot_label = '昼'
    bounds = NUTRI_BOUNDS[slot_label]
    # 具体的な代替メニュー候補（食材データ.xlsxのレシピ別栄養価から高カロリー/高たんぱく/低塩を抽出）
    hi_kcal = _nutrition_candidates(data, 'カロリー', ascending=False)
    lo_kcal = _nutrition_candidates(data, 'カロリー', ascending=True)
    hi_prot = _nutrition_candidates(data, 'たんぱく質', ascending=False)
    lo_salt = _nutrition_candidates(data, '食塩相当量', ascending=True)

    def _sug(names, verb):
        return f'{verb}（候補: {" / ".join(n[:16] for n in names)}）' if names else verb
    for month, df in data.nutrition_daily.items():
        avg = df[['kcal', 'protein', 'salt']].mean()
        label = f'{month}月(月次)'
        if avg['kcal'] < bounds['kcal'][0] or avg['kcal'] > bounds['kcal'][1]:
            low = avg['kcal'] < bounds['kcal'][0]
            pos = '下限未達' if low else '上限超過'
            sug = _sug(hi_kcal, '高カロリーのメニューに差し替え/追加') if low \
                else _sug(lo_kcal, '低カロリーのメニューに差し替え')
            viol.append({'日付': label, '曜日': slot_label, 'No': 14, 'ルール': f'エネルギー月平均が{pos}',
                         '該当箇所': f'{label}{slot_label}・月平均', '理由': f'{avg["kcal"]:.0f}kcal（基準{bounds["kcal"][0]}-{bounds["kcal"][1]}kcal）',
                         '修正提案': sug, '重要度': '高'})
        if avg['salt'] > bounds['salt_max']:
            viol.append({'日付': label, '曜日': slot_label, 'No': 14, 'ルール': '食塩相当量の月平均が上限超過',
                         '該当箇所': f'{label}{slot_label}・月平均', '理由': f'{avg["salt"]:.2f}g（上限{bounds["salt_max"]}g）',
                         '修正提案': _sug(lo_salt, '低塩分のメニューに差し替え'), '重要度': '中'})
        if avg['protein'] < bounds['protein_min']:
            viol.append({'日付': label, '曜日': slot_label, 'No': 14, 'ルール': 'たんぱく質の月平均が下限未達',
                         '該当箇所': f'{label}{slot_label}・月平均', '理由': f'{avg["protein"]:.1f}g（下限{bounds["protein_min"]}g）',
                         '修正提案': _sug(hi_prot, '高たんぱくのメニューに差し替え/追加'), '重要度': '中'})
    for month, shoku in data.nutrition_shoku.items():
        if month in data.nutrition_daily:
            continue  # 栄養価シートが優先。使用食材シートは無い月のみのフォールバック
        if not all(c in shoku.columns for c in ['カロリー', 'たんぱく質', '食塩相当量']):
            continue
        sub = shoku[shoku['isDX']]
        daily = sub.groupby('md')[['カロリー', 'たんぱく質', '食塩相当量']].sum().reset_index()
        if not len(daily):
            continue
        avg = daily[['カロリー', 'たんぱく質', '食塩相当量']].mean()
        label = f'{month}月(月次)'
        if avg['カロリー'] < bounds['kcal'][0] or avg['カロリー'] > bounds['kcal'][1]:
            low = avg['カロリー'] < bounds['kcal'][0]
            pos = '下限未達' if low else '上限超過'
            sug = _sug(hi_kcal, '高カロリーのメニューに差し替え/追加') if low \
                else _sug(lo_kcal, '低カロリーのメニューに差し替え')
            viol.append({'日付': label, '曜日': slot_label, 'No': 14, 'ルール': f'エネルギー月平均が{pos}',
                         '該当箇所': f'{label}{slot_label}・月平均', '理由': f'{avg["カロリー"]:.0f}kcal（基準{bounds["kcal"][0]}-{bounds["kcal"][1]}kcal）',
                         '修正提案': sug, '重要度': '高'})
        if avg['食塩相当量'] > bounds['salt_max']:
            viol.append({'日付': label, '曜日': slot_label, 'No': 14, 'ルール': '食塩相当量の月平均が上限超過',
                         '該当箇所': f'{label}{slot_label}・月平均', '理由': f'{avg["食塩相当量"]:.2f}g（上限{bounds["salt_max"]}g）',
                         '修正提案': _sug(lo_salt, '低塩分のメニューに差し替え'), '重要度': '中'})
        if avg['たんぱく質'] < bounds['protein_min']:
            viol.append({'日付': label, '曜日': slot_label, 'No': 14, 'ルール': 'たんぱく質の月平均が下限未達',
                         '該当箇所': f'{label}{slot_label}・月平均', '理由': f'{avg["たんぱく質"]:.1f}g（下限{bounds["protein_min"]}g）',
                         '修正提案': _sug(hi_prot, '高たんぱくのメニューに差し替え/追加'), '重要度': '中'})
    return pd.DataFrame(viol)


def check_rule15(data):
    """No.15: 週に1回、健康食材を使用する"""
    return _max_gap_check(data, is_health, 7, 15, '健康食材の使用間隔が週1回を下回る')


def check_rule17(data):
    """No.17: 1食につき赤・黄・緑の食材を必ず使用する。
    野菜マスタ_テンプレート.xlsx の「色（彩り）」列（ユーザー指定・No.9と同じマスタ）で判定する。
    「1食」＝昼は昼、夜は夜で別々に判定する（No.9と同じ考え方）。
    マスタに登録の無い野菜は検出できない点に注意（見つかり次第マスタに追記する運用）。"""
    if not data.veg_color_map:
        return pd.DataFrame()
    dr = data.date_range
    viol = []
    for d in dr:
        month = d.month
        for slot, shoku in [('昼', data.shoku.get(month)), ('夜', data.shoku_night.get(month))]:
            if shoku is None:
                continue
            md = (d.month, d.day)
            sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
            if not len(sub):
                continue
            colors = set()
            for prod in sub['商品名'].astype(str):
                colors |= veg_colors_for(prod, data.veg_color_map)
            missing = [c for c in ('赤', '黄', '緑') if c not in colors]
            if missing:
                # 代替え案はレシピ名で出す：不足している色の野菜を含むメニュー
                sugs = []
                for c in missing:
                    def _has_color(n, _c=c):
                        cols = set()
                        for p in _recipe_products(data).get(n, ()):
                            cols |= veg_colors_for(p, data.veg_color_map)
                        return _c in cols
                    cand = _recipe_replacement(data, d, ok=_has_color)
                    if cand:
                        sugs.append(f'{c}:「{cand[:22]}」')
                suggestion = ('不足色を補うメニュー候補 ' + ' / '.join(sugs)) if sugs \
                    else f'{"".join(missing)}系の食材を含むメニューを1品追加'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 17,
                    'ルール': '1食につき赤・黄・緑の食材を必ず使用', '該当箇所': f'[{slot}]',
                    '理由': f'{"".join(missing)}系の食材が0品（野菜マスタ照合）',
                    '修正提案': suggestion, '重要度': '中',
                })
    return pd.DataFrame(viol)


WEIGHT_GRAM_UNITS = {'g', 'ｇ'}
_GRAM_RE = re.compile(r'(\d+(?:\.\d+)?)\s*[gｇ]')


def _recipe_weight_g(sub):
    """1レシピ分の行（sub）から重量(g)を推定する（ユーザー指定アルゴリズム、混在パターン確認済み）。
    ・全行がg/ｇ単位（個数商材が無い＝おかず全体がg計量のレシピ）→ レシピ総量をそのまま採用
      （同一レシピ内で重複する定数のため1行分のみ。例：豚肉生姜焼き☆添えなし→70g）。
    ・個数単位（個/切/尾/枚等）の行が1つでもある場合（＝主菜が単体商材のレシピ、g行の付け合わせ野菜等が
      混在することもある）→ 個数行は「食材数量×商品名に含まれるグラム数」、g行は「食材数量(g)」をそのまま、
      それぞれ合算する（ユーザー確認済み：例 サクッとチキン アンチョビ野菜添え）。
    どの手がかりも無ければ None（重量不明として計算対象から除外）。"""
    piece_rows = sub[~sub['ユニット名'].astype(str).isin(WEIGHT_UNITS)]
    g_rows = sub[sub['ユニット名'].astype(str).isin(WEIGHT_GRAM_UNITS)]
    if not len(piece_rows):
        # 個数商材が無い＝全体がg計量のレシピ：レシピ総量を採用
        if len(g_rows):
            vals = pd.to_numeric(g_rows['レシピ総量'], errors='coerce').dropna()
            if len(vals):
                return float(vals.iloc[0])
        return None
    total = 0.0
    found = False
    for _, r in piece_rows.iterrows():
        m = _GRAM_RE.search(str(r['商品名']))
        if not m:
            continue
        qty = pd.to_numeric(r.get('食材数量'), errors='coerce')
        if pd.isna(qty):
            continue
        total += float(qty) * float(m.group(1))
        found = True
    for _, r in g_rows.iterrows():
        qty = pd.to_numeric(r.get('食材数量'), errors='coerce')
        if pd.isna(qty):
            continue
        total += float(qty)
        found = True
    return total if found else None


def check_rule18(data, min_weight_g=212):
    """No.18: 1食あたりの重量下限（現状Mサイズのみ判定：212g／Sは実データ未入手のためユーザー指定によりスキップ）。
    レシピごとに _recipe_weight_g() で重量を推定し、1食（昼/夜別）に登場する全レシピの合計と下限を比較する。
    ※容器・カップ重量は含まれていない（ルール文言上は下限に容器＋カップ重量を含む）。
    そのため本チェックは食材のみの重量であり、実際の総重量はこれより大きくなる＝本チェックは
    「これを下回れば確実にNG」という下限側の目安として運用する（容器分の上乗せは考慮できていない点に注意）。"""
    if not data.day_csv:
        return pd.DataFrame()
    viol = []
    dr = data.date_range
    for d in dr:
        for slot in ('昼', '夜'):
            df = data.day_csv.get((d.month, slot))
            if df is None:
                continue
            sub_day = df[(df['md'] == (d.month, d.day)) & (~df['レシピ名'].astype(str).str.contains('備品', na=False))]
            if not len(sub_day):
                continue
            total = 0.0
            unknown = []
            weights = []
            for recipe, grp in sub_day.groupby('レシピ名', sort=False):
                w = _recipe_weight_g(grp)
                if w is None:
                    unknown.append(str(recipe))
                    continue
                total += w
                weights.append((w, str(recipe)))
            if unknown:
                # 重量不明レシピがあれば合計は過小評価の可能性が高いため、参考情報として理由に残す
                unknown_note = f'（重量不明レシピ{len(unknown)}件を除く: ' + '/'.join(u[:10] for u in unknown[:3]) + '）'
            else:
                unknown_note = ''
            if total < min_weight_g:
                short = min_weight_g - total
                if weights:
                    w_min, r_min = min(weights)
                    sug = f'最も軽い「{r_min[:18]}」（約{w_min:.0f}g）を中心に、計{short:.0f}g分を増量'
                else:
                    sug = f'副菜等で計{short:.0f}g分を増量'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 18,
                    'ルール': '1食の重量が下限（M=212g）未達',
                    '該当箇所': f'[{slot}] 合計約{total:.0f}g',
                    '理由': f'下限{min_weight_g}gに対し約{total:.0f}g{unknown_note}（容器・カップ重量は含まず）',
                    '修正提案': sug, '重要度': '中',
                })
    return pd.DataFrame(viol)


def check_rule19(data):
    """No.19: 1食につき同じ調味料のみでの味付けはしない。
    調味料.csv（商品ID基準・No.8と同じマスタ）でその食事（昼/夜別）内に登場する調味料商品IDを集計し、
    使用されている調味料が実質1種類だけ（＝全レシピが同じ調味料でしか味付けされていない）の場合にNGとする。
    調味料が全く使われていない食事（0種類）は対象外（別問題のため）。"""
    if not data.seasoning_ids or not data.day_csv:
        return pd.DataFrame()
    viol = []
    dr = data.date_range
    for d in dr:
        for slot in ('昼', '夜'):
            df = data.day_csv.get((d.month, slot))
            if df is None:
                continue
            sub_day = df[(df['md'] == (d.month, d.day)) & (~df['レシピ名'].astype(str).str.contains('備品', na=False))]
            if not len(sub_day):
                continue
            ids_num = pd.to_numeric(sub_day['商品ID'], errors='coerce')
            season_rows = sub_day[ids_num.isin(data.seasoning_ids)]
            if not len(season_rows):
                continue
            uniq = season_rows[['商品ID', '商品名']].drop_duplicates(subset='商品ID')
            if len(uniq) == 1:
                name = str(uniq['商品名'].iloc[0])
                used_id = int(uniq['商品ID'].iloc[0])
                # 代替え案はレシピ名で出す：その調味料を使わず、別の調味料で味付けしているメニュー
                today_recipes = set(sub_day['レシピ名'].astype(str))

                def _other_seasoning(n, _uid=used_id):
                    pids = _recipe_product_ids(data).get(n, set())
                    seas = pids & data.seasoning_ids
                    return bool(seas) and _uid not in seas
                cand = _recipe_replacement(data, d, ok=_other_seasoning, exclude=today_recipes)
                suggestion = f'一部を「{cand[:26]}」等、別の調味料で味付けしたメニューに差し替え' if cand \
                    else '一部の料理を別の調味料で味付けしたメニューに差し替え'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 19,
                    'ルール': '1食につき同じ調味料のみでの味付け',
                    '該当箇所': f'[{slot}] {name[:24]}',
                    '理由': f'食事内で使われている調味料が「{name[:20]}」のみ（{len(season_rows)}レシピ行で使用）',
                    '修正提案': suggestion, '重要度': '中',
                })
    return pd.DataFrame(viol)


def is_dashi(name):
    """商品名に「だし/出汁」を含むか（NFKC正規化して判定）。
    実データ上、該当する調味料商品は「☆☆やどかり弁当　和風だし　10kg」（商品ID 1001499）の1種類のみ。"""
    n = _nfkc(name)
    return 'だし' in n or '出汁' in n


def check_rule20(data):
    """No.20: 1食につきだしの味付けを1品以上。
    調味料マスタ・実データ上「だし」を含む調味料は「☆☆やどかり弁当　和風だし」のみ確認できたため、
    商品名に「だし/出汁」を含む商材が食事（昼/夜別）内に1品も無ければNGとする（キーワード判定・マスタに専用フラグ列は無い）。"""
    if not data.day_csv:
        return pd.DataFrame()
    viol = []
    dr = data.date_range
    for d in dr:
        for slot in ('昼', '夜'):
            df = data.day_csv.get((d.month, slot))
            if df is None:
                continue
            sub_day = df[(df['md'] == (d.month, d.day)) & (~df['レシピ名'].astype(str).str.contains('備品', na=False))]
            if not len(sub_day):
                continue
            has_dashi = sub_day['商品名'].astype(str).apply(is_dashi).any()
            if not has_dashi:
                # 代替え案はレシピ名で出す：だしで味付けしているメニュー
                today_recipes = set(sub_day['レシピ名'].astype(str))
                cand = _recipe_replacement(
                    data, d,
                    ok=lambda n: any(is_dashi(p) for p in _recipe_products(data).get(n, ())),
                    exclude=today_recipes)
                sug = f'いずれかを「{cand[:26]}」等、だしで味付けしたメニューに差し替え' if cand \
                    else 'いずれかの料理をだし（和風だし）で味付けしたメニューに差し替え'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 20,
                    'ルール': '1食につきだしの味付けが0品',
                    '該当箇所': f'[{slot}]', '理由': '商品名に「だし/出汁」を含む商材が0品',
                    '修正提案': sug, '重要度': '中',
                })
    return pd.DataFrame(viol)


def check_rule21(data):
    """No.21: 禁止食材・調味料の使用禁止。
    食材データ.xlsx「禁止食材・調味料該当」シート（商品ID基準・ユーザー指定）を優先して判定する。
    day_csv/ng_product_idsが無い場合のみ、従来のキーワード判定(NG_WORDS)にフォールバックする。"""
    if data.day_csv and data.ng_product_ids:
        viol = []
        dr = data.date_range
        for d in dr:
            for slot in ('昼', '夜'):
                df = data.day_csv.get((d.month, slot))
                if df is None:
                    continue
                sub = df[(df['md'] == (d.month, d.day)) & (~df['レシピ名'].astype(str).str.contains('備品', na=False))]
                if not len(sub):
                    continue
                ids_num = pd.to_numeric(sub['商品ID'], errors='coerce')
                hit = sub[ids_num.isin(data.ng_product_ids)]
                for recipe, grp in hit.groupby('レシピ名', sort=False):
                    prods = list(dict.fromkeys(grp['商品名'].astype(str).tolist()))[:3]
                    viol.append({
                        '日付': d.strftime('%-m/%-d'), '曜日': f'{slot}/{WD_JP[d.weekday()]}', 'No': 21,
                        'ルール': '禁止食材・調味料の使用（禁止食材マスタ照合）',
                        '該当箇所': f'[{slot}] {str(recipe)[:22]} → {"/".join(prods)[:30]}',
                        '理由': '「禁止食材・調味料該当」シート登録商品を使用',
                        '修正提案': _ng_replacement(data, prods[0] if prods else None, d,
                                                    recipe_name=str(recipe)),
                        '重要度': '高',
                    })
        return pd.DataFrame(viol)
    # フォールバック：禁止食材マスタが無い場合は従来のキーワード判定
    pattern = '|'.join(NG_WORDS)
    viol = []
    md2date = {(ts.month, ts.day): ts for ts in (data.date_range if data.date_range is not None else [])}
    base_year = data.date_range[0].year if data.date_range is not None and len(data.date_range) else 2025
    for month in data.months:
        for label, shoku in [('昼', data.shoku.get(month)), ('夜', data.shoku_night.get(month))]:
            if shoku is None:
                continue
            hit = shoku[shoku['商品名'].astype(str).str.contains(pattern, na=False) |
                        shoku['レシピ名'].astype(str).str.contains(pattern, na=False)].copy()
            for (md, recipe), grp in hit.groupby(['md', 'レシピ名'], sort=False):
                if pd.isna(md):
                    continue
                m, d = md
                matched_words = set()
                matched_prods = []
                for _, r in grp.iterrows():
                    for w in NG_WORDS:
                        if w in str(r['商品名']) or w in str(recipe):
                            matched_words.add(w)
                            if w in str(r['商品名']):
                                matched_prods.append(str(r['商品名']))
                matched_prods = list(dict.fromkeys(matched_prods))[:3]
                # 違反日を実日付に解決してから代替え案を出す（履歴の「直近使用日」比較に必要）
                vdate = md2date.get((m, d))
                if vdate is None:
                    try:
                        vdate = pd.Timestamp(year=base_year, month=int(m), day=int(d))
                    except ValueError:
                        vdate = None
                sug = _ng_replacement(data, matched_prods[0] if matched_prods else None, vdate,
                                       recipe_name=str(recipe), ng_words=matched_words) \
                    if vdate is not None else '代替食材/調味料に変更'
                viol.append({
                    '日付': f'{m}/{d}', '曜日': label, 'No': 21, 'ルール': '禁止食材・調味料の使用（キーワード判定・参考）',
                    '該当箇所': f'{str(recipe)[:22]}' + (f' → {"/".join(matched_prods)[:30]}' if matched_prods else ''),
                    '理由': f'禁止ワード「{"/".join(sorted(matched_words))}」に該当',
                    '修正提案': sug, '重要度': '高',
                })
    return pd.DataFrame(viol)


def check_rule22(data):
    """No.22: 魚メニューは3日に1回（昼夕併せて）"""
    return _max_gap_check(data, is_fish, 3, 22, '魚メニューの間隔が3日を超過')


def check_rule23(data):
    """No.23: 食べにくさチェックリスト該当（参考実装）"""
    dr = data.date_range
    viol = []
    seen = set()
    for d in dr:
        names = raw_dish_names(data, d)
        for n in names:
            for kw, reason in EAT_NG.items():
                if kw == 'イカ':
                    hit = bool(re.search(r'(?<!ス)イカ', n))
                else:
                    hit = kw in n
                if hit:
                    key = (d.strftime('%-m/%-d'), n, kw)
                    if key in seen:
                        continue
                    seen.add(key)
                    safe_hist = _filtered_dish_hist(
                        data, lambda nm: not any(k in nm for k in EAT_NG))
                    cand = _pick_least_recent(safe_hist.keys(), safe_hist, d)
                    sug = f'「{cand[:20]}」等、食べにくさ該当の無いメニューに差し替え（最終判断は商品開発部）' \
                        if cand else '商品開発部のたべやすさ基準で再確認'
                    viol.append({
                        '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 23,
                        'ルール': '食べにくさチェックリスト該当（要確認）',
                        '該当箇所': f'{n[:20]}（{kw}）', '理由': f'NG例「{kw}」に該当＝{reason}',
                        '修正提案': sug, '重要度': '低（参考実装・人による最終判断が必要）',
                    })
    return pd.DataFrame(viol)


def check_rule25(data):
    """No.25: かぼちゃは週1回以上、同一曜日は4週間空ける"""
    dr = data.date_range
    kabocha_dates = []
    for d in dr:
        names = raw_dish_names(data, d)
        hit = [n for n in names if 'かぼちゃ' in n]
        if hit:
            kabocha_dates.append((d, hit[0]))
    kabocha_hist = _filtered_dish_hist(data, lambda n: 'かぼちゃ' in n)
    viol = []
    for i in range(1, len(kabocha_dates)):
        d0, _ = kabocha_dates[i - 1]
        d1, n1 = kabocha_dates[i]
        gap = (d1 - d0).days
        if gap > 7:
            cand = _pick_least_recent(kabocha_hist.keys(), kabocha_hist, d1)
            suggestion = f'間の週に「{cand[:18]}」等を追加' if cand else '間の週にかぼちゃメニューを追加'
            viol.append({
                '日付': d1.strftime('%-m/%-d'), '曜日': WD_JP[d1.weekday()], 'No': 25,
                'ルール': 'かぼちゃの使用間隔が週1回を下回る',
                '該当箇所': f'前回{d0.strftime("%-m/%-d")} → 今回{d1.strftime("%-m/%-d")}:{n1[:16]}',
                '理由': f'{gap}日間かぼちゃなし', '修正提案': suggestion, '重要度': '中',
            })
    by_weekday = {}
    for d, n in kabocha_dates:
        wd = d.weekday()
        if wd in by_weekday:
            pd0, pn0 = by_weekday[wd]
            gap = (d - pd0).days
            if gap <= 28:
                cand = _pick_least_recent(kabocha_hist.keys(), kabocha_hist, d, exclude={n})
                suggestion = f'この曜日は「{cand[:18]}」等に変更、または間隔を空ける' if cand else '曜日をずらすか間隔を空ける'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[wd], 'No': 25,
                    'ルール': 'かぼちゃが同一曜日で4週間以内に再使用',
                    '該当箇所': f'前回{pd0.strftime("%-m/%-d")}:{pn0[:14]} → 今回{d.strftime("%-m/%-d")}:{n[:14]}',
                    '理由': f'同一曜日で{gap}日しか空いていない（要29日以上）', '修正提案': suggestion, '重要度': '中',
                })
        by_weekday[wd] = (d, n)
    return pd.DataFrame(viol)


def check_rule26(data):
    """No.26: かにのふわふわは5日以上空けて使用"""
    return _min_gap_check(data, lambda n: 'かに' in n and 'ふわふわ' in n, 5, 26,
                           'かにのふわふわが5日以内に再使用', severity='低')


def _veg_rows_slot(data, date, slot):
    """指定日・指定時間帯（昼/夜）の食材行（商品ID/商品名/レシピ名）を返す（No.30専用）"""
    month = date.month
    shoku = data.shoku.get(month) if slot == '昼' else data.shoku_night.get(month)
    if shoku is None:
        return pd.DataFrame(columns=['商品ID', '商品名', 'レシピ名'])
    md = (date.month, date.day)
    sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
    sub = sub[sub['食材数量'].fillna(0) != 0]
    return sub[['商品ID', '商品名', 'レシピ名']]


def check_rule30(data):
    """No.30: 野菜使用の間隔（FDメニュールール（野菜）シート基準）。
    https://docs.google.com/spreadsheets/d/1w6ck7gAUbJIOOlDODM58QKj6nkBc2WSX5T0_Cpv7QBY (gid=1671677088)
    の内容をVEG_TIER_MASTERに反映し、商品ID単位で判定する（ユーザー確認済み）。
    ・メニュー名（レシピ名）にその食材名が明記されている場合は、必要日数を通常の2倍（doubled_days）にする。
    ・same_day_exempt=Trueの芋類/かぼちゃは、同日の昼→夜連続使用は例外的にOK（シート注記）。"""
    dr = data.date_range
    viol = []
    for match_type, key, name_kw, base_days, doubled_days, same_day_exempt, label in VEG_TIER_MASTER:
        name_kws = name_kw if isinstance(name_kw, tuple) else (name_kw,)
        occurrences = []  # (slot_datetime, date, slot, is_named, matched_product_name)
        for d in dr:
            for slot in ('昼', '夜'):
                rows = _veg_rows_slot(data, d, slot)
                if not len(rows):
                    continue
                if match_type == 'id':
                    hit = rows[pd.to_numeric(rows['商品ID'], errors='coerce') == key]
                else:
                    # 商品名（実際の食材）のみで判定する。レシピ名も含めると、同じレシピ内の
                    # 無関係な食材（例：かぼちゃ系メニューに入っているさつまいも等）まで
                    # 誤って対象商材として拾ってしまうため。
                    hit = rows[rows['商品名'].astype(str).str.contains(name_kw, na=False)]
                if not len(hit):
                    continue
                is_named = hit['レシピ名'].astype(str).apply(
                    lambda rn: any(k in rn for k in name_kws)).any()
                slot_dt = pd.Timestamp(d) + pd.Timedelta(days=(0 if slot == '昼' else 0.5))
                occurrences.append((slot_dt, d, slot, is_named, str(hit['商品名'].iloc[0])))
        for i in range(1, len(occurrences)):
            dt0, d0, slot0, _, n0 = occurrences[i - 1]
            dt1, d1, slot1, named1, n1 = occurrences[i]
            gap = (dt1 - dt0) / pd.Timedelta(days=1)
            if same_day_exempt and abs(gap - 0.5) < 1e-9:
                continue
            required = doubled_days if named1 else base_days
            if gap < required:
                # 代替え案はレシピ名で出す：当該野菜を使わず、間隔制約の緩い野菜を使うメニュー
                flex = set(VEG_FLEXIBLE_IDS)

                def _flex_recipe(n, _pid=(key if match_type == 'id' else None), _kws=name_kws):
                    if _pid is not None:
                        if _pid in _recipe_product_ids(data).get(n, set()):
                            return False
                    elif any(_recipe_has(data, n, k) for k in _kws):
                        return False
                    return bool(_recipe_product_ids(data).get(n, set()) & flex)
                cand = _recipe_replacement(data, dt1, ok=_flex_recipe)
                sug = f'「{cand[:26]}」等、間隔制約の緩い野菜を使うメニューに差し替え' if cand \
                    else '使用日をずらす'
                viol.append({
                    '日付': d1.strftime('%-m/%-d'), '曜日': f'{slot1}/{WD_JP[d1.weekday()]}', 'No': 30,
                    'ルール': f'野菜(FDメニュールール)の使用間隔違反：{label}',
                    '該当箇所': f'{n1[:24]}（前回{d0.strftime("%-m/%-d")}{slot0}）',
                    '理由': f'{gap:g}日しか空いていない（要{required}日以上{"・メニュー名記載のため2倍適用" if named1 else ""}）',
                    '修正提案': sug, '重要度': '低（参考実装）',
                })
    return pd.DataFrame(viol)


def check_rule27(data):
    """No.27: FD専用商材（魚弁当のみ、確定5商材）は平日に入れる（参考実装）＋
    「FDメニュールール」シート（★マーク商品の備考「平日夜に◯回は入れる」・ユーザー確認済みスコープ）に
    基づく、月内の平日夜 最低使用回数チェック（FD_WEEKDAY_NIGHT_QUOTA）。"""
    dr = data.date_range
    viol = []
    for d in dr:
        names = raw_dish_names(data, d)
        for n in names:
            for kw in FISH_FD_ONLY:
                if kw in n:
                    wd = d.weekday()
                    if wd >= 5:
                        nxt = _next_weekday(d)
                        sug = f'{nxt.strftime("%-m/%-d")}({WD_JP[nxt.weekday()]})等の平日枠に振り替える' \
                            if nxt is not None else '平日の枠に振り替える'
                        viol.append({
                            '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[wd], 'No': 27,
                            'ルール': 'FD専用商材（魚弁当）は平日に入れる',
                            '該当箇所': n[:30],
                            '理由': f'FD専用魚商材「{kw}」が休日（{WD_JP[wd]}）に使用されている',
                            '修正提案': sug, '重要度': '中（参考実装・魚弁当のみ）',
                        })
    # ★マーク商品の「平日夜に◯回は入れる」月内最低回数チェック（月単位の集計）。
    # 夜の食材データが無い月は、使用回数が常に0となり全★商材が誤検出になるためスキップする。
    quota_months = []
    for mth in data.months:
        sn = data.shoku_night.get(mth)
        if sn is None or not len(sn):
            data.warnings.append(
                f'{mth}月：夜の食材データが無いため、No.27の★商材「平日夜クオータ」判定をスキップしました'
                '（夜のCSVを読み込むと判定できます）')
        else:
            quota_months.append(mth)
    for pid, kw, min_count, waku in FD_WEEKDAY_NIGHT_QUOTA:
        by_month = {}
        for d in dr:
            if d.weekday() >= 5:
                continue  # 平日のみ対象
            month = d.month
            shoku = data.shoku_night.get(month)
            if shoku is None:
                continue
            md = (d.month, d.day)
            sub = shoku[(shoku['md'] == md) & (shoku['isDX'])]
            if not len(sub):
                continue
            if pid is not None:
                hit = (pd.to_numeric(sub['商品ID'], errors='coerce') == pid).any()
            else:
                hit = sub['商品名'].astype(str).str.contains(kw, na=False).any() or \
                    sub['レシピ名'].astype(str).str.contains(kw, na=False).any()
            if hit:
                by_month.setdefault(month, set()).add(d)
        for month in quota_months:
            used = sorted(by_month.get(month, set()))
            cnt = len(used)
            if cnt < min_count:
                used_txt = '／'.join(f'{u.strftime("%-m/%-d")}({WD_JP[u.weekday()]})' for u in used) or 'なし'
                viol.append({
                    '日付': f'{month}月(月次)', '曜日': '夜', 'No': 27,
                    'ルール': f'★商材の平日夜クオータ未達：{waku}「{kw}」',
                    '該当箇所': f'{month}月の平日夜 全体（使用日: {used_txt}）',
                    '理由': f'平日夜の使用が{cnt}回のみ（月{min_count}回以上必要）※日単位ではなく月単位の集計',
                    '修正提案': f'平日夜の枠に「{kw}」をあと{min_count - cnt}回追加する', '重要度': '中（FDメニュールール準拠）',
                })
    return pd.DataFrame(viol)


def check_rule28(data):
    """No.28: 本日の魚料理は平日の夜に採用する"""
    viol = []
    for (d, wd, slot, pos, name) in data.rows:
        if '本日の魚料理' in name:
            problems = []
            if slot != '夜':
                problems.append(f'{slot}に使用（要:夜）')
            if wd in ('土', '日'):
                problems.append(f'{wd}曜（休日）に使用（要:平日）')
            if problems:
                nxt = _next_weekday(d) if wd in ('土', '日') else None
                sug = f'{nxt.strftime("%-m/%-d")}({WD_JP[nxt.weekday()]})等の平日の夜枠に振り替える' \
                    if nxt is not None else '同日の夜枠に移す（平日夜が要件）'
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': wd, 'No': 28,
                    'ルール': '本日の魚料理は平日の夜に採用する',
                    '該当箇所': name[:30],
                    '理由': ' / '.join(problems),
                    '修正提案': sug, '重要度': '中',
                })
    return pd.DataFrame(viol)


def check_rule29(data):
    """No.29: おまかせメニューを昼・夜月2回以上採用"""
    omakase_count = Counter()
    seen_months = set()
    for (d, wd, slot, pos, name) in data.rows:
        seen_months.add(d.month)
        if 'おまかせ' in name:
            omakase_count[(d.month, slot)] += 1
    omakase_names = sorted({name for (d, wd, slot, pos, name) in data.rows if 'おまかせ' in name},
                            key=lambda n: -len(n))
    viol = []
    for month in sorted(seen_months):
        for slot in ['昼', '夜']:
            cnt = omakase_count.get((month, slot), 0)
            if cnt < 2:
                example = f'（例:「{omakase_names[0][:18]}」等）' if omakase_names else ''
                viol.append({
                    '日付': f'{month}月(月次)', '曜日': '-', 'No': 29,
                    'ルール': 'おまかせメニューを昼・夜月2回以上採用',
                    '該当箇所': f'{month}月{slot}',
                    '理由': f'おまかせメニューが{cnt}回のみ（月2回以上必要）',
                    '修正提案': f'{slot}のおまかせ枠を追加する{example}', '重要度': '中',
                })
    return pd.DataFrame(viol)


def check_rule24(data):
    """No.24: 白和えは副菜に分類し、サラダ使用時は酢を混ぜる"""
    viol = []
    for (d, wd, slot, pos, name) in data.rows:
        if '白和え' in name:
            if pos == 'サラダ' and '酢' not in name:
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': wd, 'No': 24,
                    'ルール': '白和えは副菜に分類し、サラダ使用時は酢を混ぜる',
                    '該当箇所': f'{slot}:{pos}:{name[:20]}',
                    '理由': 'サラダ枠での白和え使用だが「酢」の明記がない',
                    '修正提案': '酢を加える、または副菜枠に変更する', '重要度': '低',
                })
            elif pos not in ('副菜1', '副菜2', 'サラダ'):
                viol.append({
                    '日付': d.strftime('%-m/%-d'), '曜日': wd, 'No': 24,
                    'ルール': '白和えは副菜に分類し、サラダ使用時は酢を混ぜる',
                    '該当箇所': f'{slot}:{pos}:{name[:20]}',
                    '理由': f'白和えが副菜以外（{pos}）で使用されている',
                    '修正提案': '副菜枠に変更する', '重要度': '低',
                })
    return pd.DataFrame(viol)


def check_rule31(data):
    """No.31: マッシュ系調理、メニュー名に明記があれば同日2品以上NG"""
    dr = data.date_range
    viol = []
    for d in dr:
        names = raw_dish_names(data, d)
        mash = sorted(n for n in names if 'マッシュ' in n)
        if len(mash) >= 2:
            nomash_hist = _filtered_dish_hist(data, lambda nm: 'マッシュ' not in nm)
            cand = _pick_least_recent(nomash_hist.keys(), nomash_hist, d)
            sug = f'一方を「{cand[:20]}」等に差し替え、またはメニュー名から「マッシュ」を外す' \
                if cand else '一方を別の調理法名に'
            viol.append({
                '日付': d.strftime('%-m/%-d'), '曜日': WD_JP[d.weekday()], 'No': 31,
                'ルール': 'メニュー名に「マッシュ」と明記された商材が同日2品以上',
                '該当箇所': ' / '.join(mash), '理由': f'マッシュ系が{len(mash)}品',
                '修正提案': sug, '重要度': '低',
            })
    return pd.DataFrame(viol)


ALL_RULES = [
    ('No.1 メイン/サブ商材1週間ルール（酷似商材含む）', check_rule1),
    # ('No.2 メイン/サブ見た目酷似1週間+月2回ルール', check_rule2),  # ユーザー指示によりいったん無効化（商品IDが異なる場合は検出しない方針）
    ('No.3/5 挽肉・鶏豚牛のメイン/サブ重複', check_rule3_5),
    ('No.4/36 コロッケ間隔', check_rule4_36),
    ('No.6 同一食材の複数レシピ重複（1食内）', check_rule6),
    ('No.8 食材+調味料の複数レシピ重複（1食内）', check_rule8),
    ('No.9 見た目（色）が同じ野菜の2日連続', check_rule9),
    ('No.10 単一食材メニューの1週間ルール', check_rule10),
    ('No.11 自然解凍1品ルール', check_rule11),
    ('No.7 大豆系商材の同日重複', check_rule7),
    ('No.12 揚げ物3品まで', check_rule12),
    ('No.14 栄養素基準（月平均）', check_rule14),
    ('No.15 健康食材 週1回以上', check_rule15),
    ('No.17 1食で赤・黄・緑を使用', check_rule17),
    ('No.18 1食の重量下限（M=212g）', check_rule18),
    ('No.19 同じ調味料のみでの味付け禁止', check_rule19),
    ('No.20 だし味付け1品以上', check_rule20),
    ('No.21 禁止食材・調味料', check_rule21),
    ('No.22 魚メニュー3日に1回', check_rule22),
    # ('No.23 食べにくさチェックリスト', check_rule23),  # ユーザー指示によりチェック対象外
    ('No.24 白和えの分類', check_rule24),
    ('No.25 かぼちゃ週1回・同曜日4週間', check_rule25),
    ('No.26 かにのふわふわ5日以上空ける', check_rule26),
    ('No.27 FD専用商材（魚弁当）は平日', check_rule27),
    ('No.28 本日の魚料理は平日の夜', check_rule28),
    ('No.29 おまかせメニュー月2回以上', check_rule29),
    ('No.30 野菜使用間隔（FDメニュールール）', check_rule30),
    ('No.31 マッシュ系同日重複', check_rule31),
]


def run_all_checks(xlsx_path, night_csv_paths=None, day_csv_paths=None, veg_master_path=None,
                    seasoning_csv_path=None, fried_master_path=None, ai_client=None):
    data = load_workbook_data(xlsx_path, night_csv_paths, day_csv_paths, veg_master_path,
                               seasoning_csv_path, fried_master_path)
    data.ai_client = ai_client
    frames = []
    for label, fn in ALL_RULES:
        try:
            result = fn(data)
        except Exception as e:
            data.warnings.append(f'{label} の判定中にエラーが発生しスキップしました（{e}）')
            continue
        if isinstance(result, tuple):
            for r in result:
                if len(r):
                    frames.append(r.reindex(columns=cols_std))
        else:
            if len(result):
                frames.append(result.reindex(columns=cols_std))
    if frames:
        combined = pd.concat(frames, ignore_index=True)
        combined['No'] = combined['No'].astype(int)
        # 「曜日」を「昼夜」に置き換え、日付順（同日内は No 順）に並べる（ユーザー指定）
        combined['昼夜'] = combined.apply(_derive_slot, axis=1)
        combined['_k'] = combined['日付'].map(_date_sort_key)
        slot_order = {'昼': 0, '昼夜': 1, '夜': 2, '-': 3}
        combined['_s'] = combined['昼夜'].map(lambda x: slot_order.get(x, 9))
        combined = (combined.sort_values(['_k', '_s', 'No'])
                            .drop(columns=['_k', '_s', '曜日'])
                            .reindex(columns=OUT_COLS)
                            .reset_index(drop=True))
    else:
        combined = pd.DataFrame(columns=OUT_COLS)
    n_days = len(set(r[0] for r in data.rows)) if data.rows else 0
    summary = {
        'months': data.months,
        'n_days': n_days,
        'total': len(combined),
        'by_rule': combined['No'].value_counts().sort_index().to_dict() if len(combined) else {},
        'warnings': data.warnings,
    }
    return combined, summary


RULE_NAME_JP = {
    1: 'メイン/サブ商材1週間', 2: '見た目酷似1週間+月2回', 3: '挽肉のメイン/サブ重複',
    4: 'コロッケ間隔', 5: '鶏豚牛のメイン/サブ重複', 6: '同一食材の複数レシピ重複',
    7: '大豆系は半日空ける', 8: '食材と調味料の中身被り', 9: '同じ色の野菜が2日連続',
    10: '単一食材メニュー1週間', 11: '自然解凍1品', 12: '当日揚げ3品まで',
    13: '盛付工数3工程まで', 14: '栄養素基準（月平均）', 15: '健康食材 週1回以上',
    16: '固形は2種まで', 17: '赤・黄・緑を使用', 18: '1食の重量下限',
    19: '同じ調味料のみの味付け禁止', 20: 'だし味付け1品以上', 21: '禁止食材・調味料',
    22: '魚メニュー3日に1回', 23: '食べにくさチェック', 24: '白和えの分類',
    25: 'かぼちゃ週1回・同曜日4週間', 26: 'かにのふわふわ5日以上', 27: 'FD専用商材・★平日夜クオータ',
    28: '本日の魚料理は平日の夜', 29: 'おまかせ月2回以上', 30: '野菜の使用間隔',
    31: 'マッシュ系同日重複', 36: 'コロッケとクリームコロッケ',
}


def write_report(combined, n_days, out_path):
    """違反候補リスト／サマリー／チェックフローの3シートで出力する。
    ・「曜日」ではなく「昼夜」列を出す（ユーザー指定）。
    ・日付順（月次集計行はその月の先頭）に並べ、同日内は重要度順にする。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    out = combined.copy() if len(combined) else pd.DataFrame(columns=OUT_COLS)
    if len(out):
        # 日付順 → 同日内は 昼 → 昼夜 → 夜 → No順（重要度では並べ替えない。重要度は行の色で表現する）
        slot_order = {'昼': 0, '昼夜': 1, '夜': 2, '-': 3}
        out['_k'] = out['日付'].map(_date_sort_key)
        out['_s'] = out['昼夜'].map(lambda x: slot_order.get(x, 9)) if '昼夜' in out.columns else 0
        out = out.sort_values(['_k', '_s', 'No']).drop(columns=['_k', '_s']).reset_index(drop=True)
    wb = Workbook()
    H = Font(name='Meiryo', bold=True, color='FFFFFF', size=10)
    HF = PatternFill('solid', fgColor='C0703B')
    SUB = Font(name='Meiryo', bold=True, size=11, color='C0703B')
    BODY = Font(name='Meiryo', size=10)
    sevfill = {'高': PatternFill('solid', fgColor='F4CCCC'), '中': PatternFill('solid', fgColor='FFF2CC')}
    thin = Side(style='thin', color='D9C9B5')
    BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, n, row=1):
        for c in range(1, n + 1):
            cc = ws.cell(row, c)
            cc.font = H
            cc.fill = HF
            cc.border = BORD
            cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws = wb.active
    ws.title = '違反候補リスト'
    ws.merge_cells('A1:H1')
    ws['A1'] = '■ お弁当メニュー違反チェック結果（食材チェックAI）'
    ws['A1'].font = Font(name='Meiryo', bold=True, size=13, color='C0703B')
    ws.append([])
    ws.append(OUT_COLS)
    hdr(ws, len(OUT_COLS), row=3)
    for _, r in out.iterrows():
        ws.append([r.get(c) for c in OUT_COLS])
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        sev = str(row[7].value)[:1]
        for cell in row:
            cell.font = BODY
            cell.border = BORD
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.fill = sevfill.get(sev, PatternFill())
    if ws.max_row < 4:
        ws.append(['(違反候補なし)'])
    ws.freeze_panes = 'A4'
    for col, w in zip('ABCDEFGH', [8, 6, 5, 24, 44, 30, 34, 10]):
        ws.column_dimensions[col].width = w
    ws2 = wb.create_sheet('サマリー')
    ws2['A1'] = '検出サマリー'
    ws2['A1'].font = SUB
    ws2.append([])
    ws2.append(['ルール', '検出件数', '重要度内訳'])
    hdr(ws2, 3, row=3)
    if len(out):
        for no in sorted(out['No'].unique()):
            s = out[out['No'] == no]
            sv = '/'.join(f'{k}{v}' for k, v in s['重要度'].value_counts().items())
            ws2.append([f'No.{no} {RULE_NAME_JP.get(no, "")}', len(s), sv])
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row):
        for cell in row:
            cell.font = BODY
            cell.border = BORD
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws2.append([])
    ws2.append(['合計', len(out), f'検査日数 {n_days}日'])
    ws2.cell(ws2.max_row, 1).font = Font(name='Meiryo', bold=True, size=10)
    for col, w in zip('ABC', [30, 12, 28]):
        ws2.column_dimensions[col].width = w
    ws3 = wb.create_sheet('チェックフロー')
    ws3['A1'] = '食材チェックAI 処理フロー'
    ws3['A1'].font = Font(name='Meiryo', bold=True, size=13, color='C0703B')
    ws3.append([])
    flow = [('① ルール・マスタ内蔵', '構成ルールと、野菜マスタ(色)/調味料マスタ/当日揚げ・禁止食材マスタ/FDメニュールールを保持'),
            ('② 入力', 'メニューワークブック＋食材CSV（昼/夜）＋各マスタファイル'),
            ('③ 展開・紐づけ', '日×昼夜×5ポジションに分解→商品ID単位で食材を紐づけ→マスタで色/系統/調理法に変換'),
            ('④ ルール適用', '基礎調味料・基礎野菜は除外、枠タイトル付きメニュー名は名寄せ、備品/水は除外'),
            ('⑤ 出力', '違反候補のみを日付順に抽出（日付・昼夜・ルール・該当箇所・理由・代替え案・重要度）')]
    for a, b in flow:
        ws3.append([a, b])
        for c in (1, 2):
            ws3.cell(ws3.max_row, c).font = BODY
            ws3.cell(ws3.max_row, c).alignment = Alignment(vertical='top', wrap_text=True)
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 78
    wb.save(out_path)
